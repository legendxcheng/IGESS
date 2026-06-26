import json
import subprocess
import sys
from pathlib import Path
from shutil import copytree

from openpyxl import load_workbook

from igess.luban_exporter import export_registered_workbooks
from igess.loader import ConfigLoader


DATAS = Path("data-tables/Datas")
EXPORTS = Path("examples/shelldiver_v0/luban_exports")


def test_luban_source_workbooks_have_marker_rows():
    for path in sorted(DATAS.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active

        assert sheet["A1"].value == "##var", path
        assert sheet["A2"].value == "##", path
        assert sheet["A3"].value == "##type", path
        assert sheet["B1"].value not in (None, ""), path
        assert sheet["B3"].value not in (None, ""), path


def test_luban_registry_lists_runtime_tables():
    workbook = load_workbook(DATAS / "__tables__.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    registered = {row[1] for row in sheet.iter_rows(min_row=4, values_only=True)}

    assert {
        "resources",
        "generators",
        "upgrades",
        "constants",
        "milestones",
        "prestige_layers",
    }.issubset(registered)


def test_resources_workbook_has_populated_dimension_column():
    workbook = load_workbook(DATAS / "resources.xlsx", read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    dimension_index = headers.index("dimension") + 1

    dimensions = [
        sheet.cell(row=row, column=dimension_index).value
        for row in range(4, sheet.max_row + 1)
    ]

    assert dimensions
    assert all(str(value).strip() for value in dimensions)


def test_luban_export_rows_carry_source_metadata():
    raw = ConfigLoader.load(
        "examples/shelldiver_v0/economy.yaml",
        "examples/shelldiver_v0/luban_exports",
    )

    fisherman = next(row for row in raw.generators if row.id == "fisherman")

    assert fisherman.source_ref is not None
    assert fisherman.source_ref.table == "generators"
    assert fisherman.source_ref.workbook == "generators.xlsx"
    assert fisherman.source_ref.row == 4


def test_sample_json_exports_include_explicit_source_metadata():
    for path in sorted(EXPORTS.glob("*.json")):
        rows = json.loads(path.read_text(encoding="utf-8"))
        assert rows, path
        for row in rows:
            source = row.get("_source")
            assert source is not None, (path, row["id"])
            assert source["table"] == path.stem
            assert source["workbook"] == f"{path.stem}.xlsx"
            assert isinstance(source["row"], int)


def test_loader_trusts_explicit_source_metadata(tmp_path):
    tables = tmp_path / "tables"
    copytree(EXPORTS, tables)
    generator_path = tables / "generators.json"
    generators = json.loads(generator_path.read_text(encoding="utf-8"))
    generators[0]["_source"] = {
        "table": "generators",
        "workbook": "generators.xlsx",
        "row": 44,
    }
    generator_path.write_text(
        json.dumps(generators, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )

    raw = ConfigLoader.load("examples/shelldiver_v0/economy.yaml", tables)
    fisherman = next(row for row in raw.generators if row.id == "fisherman")

    assert fisherman.source_ref is not None
    assert fisherman.source_ref.row == 44


def test_export_registered_workbooks_matches_checked_in_json(tmp_path):
    generated = tmp_path / "exports"

    export_registered_workbooks(DATAS, generated)

    for expected_path in sorted(EXPORTS.glob("*.json")):
        generated_path = generated / expected_path.name
        assert generated_path.read_bytes() == expected_path.read_bytes()


def test_export_registered_workbooks_removes_stale_json(tmp_path):
    generated = tmp_path / "exports"
    generated.mkdir()
    stale = generated / "stale_table.json"
    stale.write_text("[]\n", encoding="utf-8")

    export_registered_workbooks(DATAS, generated)

    assert not stale.exists()
    assert sorted(path.name for path in generated.glob("*.json")) == sorted(
        path.name for path in EXPORTS.glob("*.json")
    )


def test_export_registered_workbooks_rejects_unsupported_registry_mode(tmp_path):
    datas = tmp_path / "Datas"
    copytree(DATAS, datas)
    workbook = load_workbook(datas / "__tables__.xlsx")
    sheet = workbook.active
    sheet["D4"] = "list"
    workbook.save(datas / "__tables__.xlsx")

    try:
        export_registered_workbooks(datas, tmp_path / "exports")
    except ValueError as exc:
        assert "unsupported mode" in str(exc)
    else:
        raise AssertionError("Expected unsupported registry mode to fail")


def test_export_registered_workbooks_rejects_unsupported_registry_key(tmp_path):
    datas = tmp_path / "Datas"
    copytree(DATAS, datas)
    workbook = load_workbook(datas / "__tables__.xlsx")
    sheet = workbook.active
    sheet["E4"] = "name"
    workbook.save(datas / "__tables__.xlsx")

    try:
        export_registered_workbooks(datas, tmp_path / "exports")
    except ValueError as exc:
        assert "unsupported key" in str(exc)
    else:
        raise AssertionError("Expected unsupported registry key to fail")


def test_cli_export_tables_matches_checked_in_json(tmp_path):
    generated = tmp_path / "exports"

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "igess.cli",
            "export-tables",
            "--datas",
            str(DATAS),
            "--out",
            str(generated),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    assert "Exported 6 tables" in result.stdout
    for expected_path in sorted(EXPORTS.glob("*.json")):
        generated_path = generated / expected_path.name
        assert generated_path.read_bytes() == expected_path.read_bytes()
