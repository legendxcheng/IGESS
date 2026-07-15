from __future__ import annotations

from dataclasses import FrozenInstanceError
import hashlib
import json
from collections.abc import Iterator, Mapping
from contextlib import contextmanager
import os
from pathlib import Path
import shutil
import stat
import subprocess
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from igess.authoring import ModelChange
from igess.authoring import probe as probe_module
from igess.authoring import status as status_module
from igess.authoring.entity_schema import get_entity_schema
from igess.authoring.probe import EligibilityFinding
from igess.authoring.probe import TenTickProbeResult
from igess.authoring.status import ModelStatus, derive_status
from igess.authoring.templates import initialize_authoring_project
from igess.authoring.project import AuthoringProject
from igess.authoring.response import AuthoringError
from igess.authoring.workbook_source import upsert_workbook_entity
from igess.authoring.yaml_source import upsert_yaml_entity
from igess.schema import Event, SimulationResult, TimelineRow


def _blank_project(tmp_path: Path) -> AuthoringProject:
    root = initialize_authoring_project(tmp_path / "project", "status_test")
    return AuthoringProject.discover(root)


def _upsert(
    project: AuthoringProject,
    entity: str,
    entity_id: str,
    fields: dict[str, object],
) -> None:
    change = ModelChange(1, "upsert", entity, entity_id, fields)
    schema = get_entity_schema(entity)
    if schema.storage_kind == "workbook":
        upsert_workbook_entity(project.datas / schema.storage_name, change)
    else:
        upsert_yaml_entity(project.config, change)


def _add_resource(project: AuthoringProject) -> None:
    _upsert(project, "resource", "gold", {"name": "Gold", "dimension": "currency"})


def _add_activity_route(project: AuthoringProject) -> None:
    _add_resource(project)
    _upsert(
        project,
        "activity",
        "gather",
        {"name": "Gather", "source_type": "active", "unlock_condition": "always"},
    )
    _upsert(
        project,
        "activity_output",
        "gather_gold",
        {
            "activity_id": "gather",
            "output_resource": "gold",
            "amount_per_second": "1",
        },
    )
    _upsert(
        project,
        "player_profile",
        "default",
        {
            "source_efficiency": {
                "active": "1",
                "generator": "1",
                "offline": "1",
                "milestone": "1",
                "prestige": "1",
            },
            "behavior_policy": "cheap_unlock_first",
            "session_pattern": "authoring_default",
            "prestige_policy": "conservative",
            "activity_weights": {"gather": "1"},
            "luck": "1",
        },
    )


def _add_generator_route(
    project: AuthoringProject,
    *,
    base_cost: str = "0",
    starting_gold: str | None = None,
) -> None:
    _add_resource(project)
    _upsert(
        project,
        "generator",
        "mine",
        {
            "name": "Mine",
            "generator_type": "building",
            "output_resource": "gold",
            "source_type": "generator",
            "base_output": "1",
            "base_cost": base_cost,
            "cost_resource": "gold",
            "cost_growth": "1.15",
            "unlock_condition": "always",
        },
    )
    if starting_gold is not None:
        _upsert(project, "constant", "starting_gold", {"value": starting_gold})


def _add_formal_scenario(project: AuthoringProject) -> None:
    _upsert(
        project,
        "scenario",
        "formal",
        {
            "duration_hours": "1",
            "time_mode": "tick",
            "profiles": ["default"],
            "start_state": "new_player",
            "record_interval_seconds": 60,
            "outputs": ["resource_curve"],
        },
    )


def _break_registry(project: AuthoringProject) -> None:
    registry = project.datas / "__tables__.xlsx"
    workbook = load_workbook(registry)
    workbook.active["A1"] = "bad-marker"
    workbook.save(registry)
    workbook.close()


def _write_sentinel_resource_workbook(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(source.read_bytes())
    workbook = load_workbook(target)
    workbook.active.append([None, "outside_sentinel", "Outside", "currency"])
    workbook.save(target)
    workbook.close()


def _manifest(root: Path) -> dict[str, tuple[str, int, str]]:
    result: dict[str, tuple[str, int, str]] = {}
    pending = [root]
    while pending:
        directory = pending.pop()
        with os.scandir(directory) as entries:
            selected = sorted(entries, key=lambda item: item.name)
        for entry in selected:
            path = Path(entry.path)
            relative = path.relative_to(root).as_posix()
            identity = entry.stat(follow_symlinks=False)
            mode = identity.st_mode
            indirection = _test_path_is_indirection(path, identity)
            if stat.S_ISDIR(mode) and not indirection:
                result[relative] = ("dir", mode, "")
                pending.append(path)
            elif stat.S_ISREG(mode):
                result[relative] = (
                    "file",
                    mode,
                    hashlib.sha256(path.read_bytes()).hexdigest(),
                )
            else:
                result[relative] = ("indirection" if indirection else "other", mode, "")
    return result


def _test_path_is_indirection(path: Path, identity: os.stat_result) -> bool:
    if stat.S_ISLNK(identity.st_mode):
        return True
    is_junction = getattr(path, "is_junction", None)
    if callable(is_junction) and is_junction():
        return True
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    return bool(getattr(identity, "st_file_attributes", 0) & reparse_flag)


def _classifier_simulation(category: str | None) -> SimulationResult:
    rows: list[TimelineRow] = []
    for time_seconds in range(11):
        final = time_seconds == 10
        rows.append(
            TimelineRow(
                scenario_id="smoke",
                profile_id="default",
                time_seconds=time_seconds,
                resources={
                    "gold": (
                        "0.0000000000000000000000000001"
                        if final and category == "resource_value"
                        else ("0e99" if final and category is None else "0")
                    )
                },
                generators_owned={
                    "mine": 1 if final and category == "owned_generator_count" else 0
                },
                upgrades_purchased=(
                    ["boost"] if final and category == "purchased_upgrade_set" else []
                ),
                total_cps="999" if final else "0",
                prestige_counts={
                    "rebirth": 1 if final and category == "prestige_count" else 0
                },
            )
        )
    events = (
        [Event("smoke", "default", 10, "unlock_generator", "mine", {})]
        if category is None
        else []
    )
    return SimulationResult("smoke", rows, events)


def test_model_status_payload_is_frozen_json_safe_and_defensive() -> None:
    issue = EligibilityFinding("missing", "Add a rule", "resource", "gold")
    source_counts = {"resource": 1}
    status = ModelStatus(
        model_digest="sha256:" + "a" * 64,
        structural_valid=True,
        smoke_eligible=False,
        state="incomplete",
        entity_counts=source_counts,
        missing_requirements=(issue,),
        warnings=(),
        available_scenarios=("smoke",),
        latest_smoke_run_id=None,
    )
    source_counts["resource"] = 99

    expected = {
        "model_digest": "sha256:" + "a" * 64,
        "structural_valid": True,
        "smoke_eligible": False,
        "state": "incomplete",
        "entity_counts": {"resource": 1},
        "missing_requirements": [issue.to_payload()],
        "warnings": [],
        "available_scenarios": ["smoke"],
        "latest_smoke_run_id": None,
    }
    assert status.to_payload() == expected
    json.dumps(status.to_payload())
    payload = status.to_payload()
    payload["entity_counts"]["resource"] = 7
    payload["missing_requirements"].clear()
    assert status.to_payload() == expected
    with pytest.raises(TypeError):
        status.entity_counts["resource"] = 2  # type: ignore[index]
    with pytest.raises(FrozenInstanceError):
        status.state = "failed"  # type: ignore[misc]


@pytest.mark.parametrize(
    ("state", "structural_valid", "smoke_eligible", "scenarios"),
    [
        ("failed", False, False, ("formal",)),
        ("incomplete", True, False, ("formal",)),
        ("incomplete", True, True, ("smoke",)),
        ("incomplete", True, True, ("formal", "smoke")),
        ("runnable", True, True, ("smoke",)),
        ("ready", True, True, ("formal", "smoke")),
    ],
)
def test_model_status_accepts_exact_valid_state_matrix(
    state: str,
    structural_valid: bool,
    smoke_eligible: bool,
    scenarios: tuple[str, ...],
) -> None:
    status = ModelStatus(
        model_digest="sha256:" + "a" * 64,
        structural_valid=structural_valid,
        smoke_eligible=smoke_eligible,
        state=state,  # type: ignore[arg-type]
        entity_counts={"resource": 0},
        missing_requirements=(),
        warnings=(),
        available_scenarios=scenarios,
        latest_smoke_run_id="prior-smoke-1",
    )

    payload = status.to_payload()
    assert payload["state"] == state
    assert payload["structural_valid"] is structural_valid
    assert payload["smoke_eligible"] is smoke_eligible
    assert payload["entity_counts"] == {"resource": 0}
    assert payload["available_scenarios"] == list(scenarios)
    json.dumps(payload)


@pytest.mark.parametrize(
    ("state", "structural_valid", "smoke_eligible"),
    [
        (state, structural_valid, smoke_eligible)
        for state in ("failed", "incomplete", "runnable", "ready")
        for structural_valid in (False, True)
        for smoke_eligible in (False, True)
        if (state, structural_valid, smoke_eligible)
        not in {
            ("failed", False, False),
            ("incomplete", True, False),
            ("incomplete", True, True),
            ("runnable", True, True),
            ("ready", True, True),
        }
    ],
)
def test_model_status_rejects_every_contradictory_state_combination(
    state: str,
    structural_valid: bool,
    smoke_eligible: bool,
) -> None:
    scenarios = (
        ("formal", "smoke")
        if state == "ready"
        else (("smoke",) if smoke_eligible else ())
    )
    with pytest.raises(ValueError):
        ModelStatus(
            model_digest="sha256:" + "a" * 64,
            structural_valid=structural_valid,
            smoke_eligible=smoke_eligible,
            state=state,  # type: ignore[arg-type]
            entity_counts={},
            available_scenarios=scenarios,
        )


@pytest.mark.parametrize(
    ("state", "smoke_eligible", "scenarios"),
    [
        ("ready", True, ()),
        ("ready", True, ("smoke",)),
        ("ready", True, ("formal",)),
        ("runnable", True, ()),
        ("runnable", True, ("formal",)),
        ("runnable", True, ("formal", "smoke")),
        ("incomplete", True, ()),
        ("incomplete", True, ("formal",)),
    ],
)
def test_model_status_rejects_scenario_state_contradictions(
    state: str,
    smoke_eligible: bool,
    scenarios: tuple[str, ...],
) -> None:
    with pytest.raises(ValueError):
        ModelStatus(
            model_digest="sha256:" + "a" * 64,
            structural_valid=True,
            smoke_eligible=smoke_eligible,
            state=state,  # type: ignore[arg-type]
            entity_counts={},
            available_scenarios=scenarios,
        )


def test_model_status_normalizes_scenarios_sorted_unique_and_defensively() -> None:
    source = ["zeta", "smoke", "alpha", "zeta"]
    status = ModelStatus(
        model_digest="sha256:" + "a" * 64,
        structural_valid=True,
        smoke_eligible=True,
        state="incomplete",
        entity_counts={},
        available_scenarios=source,  # type: ignore[arg-type]
    )
    source.clear()

    assert status.available_scenarios == ("alpha", "smoke", "zeta")
    payload = status.to_payload()
    assert payload["available_scenarios"] == ["alpha", "smoke", "zeta"]
    payload["available_scenarios"].clear()
    assert status.to_payload()["available_scenarios"] == ["alpha", "smoke", "zeta"]


def test_blank_project_derives_complete_incomplete_payload(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)

    status = derive_status(project, lambda: None)

    assert status.structural_valid is True
    assert status.smoke_eligible is False
    assert status.state == "incomplete"
    assert tuple(status.entity_counts) == (
        "resource",
        "generator",
        "activity",
        "activity_output",
        "upgrade",
        "constant",
        "milestone",
        "prestige_layer",
        "formula",
        "generator_type",
        "source_type",
        "modifier_type",
        "behavior_policy",
        "session_pattern",
        "player_profile",
        "scenario",
        "rng_table",
        "rng_scenario",
        "regression_gate",
    )
    assert status.entity_counts["resource"] == 0
    assert status.entity_counts["formula"] == 3
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id is None
    assert {item.code for item in status.missing_requirements} >= {
        "no_resources",
        "no_executable_behavior",
    }
    assert [warning.code for warning in status.warnings] == ["exports_stale"]


def test_resource_only_uses_current_workbook_not_stale_committed_json(
    tmp_path: Path,
) -> None:
    project = _blank_project(tmp_path)
    (project.exports / "resources.json").write_text("[]\n", encoding="utf-8")
    old_export = (project.exports / "resources.json").read_bytes()
    _add_resource(project)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "incomplete"
    assert status.structural_valid
    assert status.entity_counts["resource"] == 1
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert [warning.code for warning in status.warnings] == ["exports_stale"]
    assert (project.exports / "resources.json").read_bytes() == old_export
    assert _manifest(project.root) == before


def test_real_activity_route_progresses_to_runnable_then_ready(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)

    runnable = derive_status(project, lambda: None)

    assert runnable.state == "runnable"
    assert runnable.structural_valid and runnable.smoke_eligible
    assert runnable.missing_requirements == ()
    assert runnable.entity_counts["activity"] == 1
    assert runnable.entity_counts["activity_output"] == 1

    _upsert(
        project,
        "scenario",
        "formal",
        {
            "duration_hours": "1",
            "time_mode": "tick",
            "profiles": ["default"],
            "start_state": "new_player",
            "record_interval_seconds": 60,
            "outputs": ["resource_curve"],
        },
    )
    ready = derive_status(project, lambda: SimpleNamespace(run_id="prior-smoke-1"))

    assert ready.state == "ready"
    assert ready.available_scenarios == ("formal", "smoke")
    assert ready.latest_smoke_run_id == "prior-smoke-1"


def test_generator_requires_affordability_and_every_smoke_profile_efficiency(
    tmp_path: Path,
) -> None:
    project = _blank_project(tmp_path)
    _add_generator_route(project, base_cost="10")
    _upsert(
        project,
        "player_profile",
        "slow",
        {
            "source_efficiency": {"generator": "0"},
            "behavior_policy": "cheap_unlock_first",
            "session_pattern": "authoring_default",
            "prestige_policy": "conservative",
            "activity_weights": {},
            "luck": "1",
        },
    )
    _upsert(
        project,
        "scenario",
        "smoke",
        {
            "duration_hours": "0.002777777777777778",
            "time_mode": "tick",
            "profiles": ["default", "slow"],
            "start_state": "new_player",
            "record_interval_seconds": 1,
            "outputs": ["resource_curve"],
        },
    )

    blocked = derive_status(project, lambda: None)

    assert blocked.state == "incomplete"
    assert {item.code for item in blocked.missing_requirements} >= {
        "generator_efficiency_nonpositive",
        "generator_unaffordable",
    }

    _upsert(project, "constant", "starting_gold", {"value": "10"})
    _upsert(
        project,
        "player_profile",
        "slow",
        {
            "source_efficiency": {"generator": "0.5"},
            "behavior_policy": "cheap_unlock_first",
            "session_pattern": "authoring_default",
            "prestige_policy": "conservative",
            "activity_weights": {},
            "luck": "1",
        },
    )

    runnable = derive_status(project, lambda: None)

    assert runnable.state == "runnable"
    assert runnable.smoke_eligible


def test_no_change_probe_remains_incomplete_not_failed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    monkeypatch.setattr(
        status_module,
        "run_ten_tick_probe",
        lambda _model: TenTickProbeResult(
            False,
            (EligibilityFinding("smoke_no_state_change", "No observable change."),),
        ),
    )

    status = derive_status(project, lambda: None)

    assert status.state == "incomplete"
    assert status.structural_valid and status.smoke_eligible
    assert [finding.code for finding in status.missing_requirements] == [
        "smoke_no_state_change"
    ]


@pytest.mark.parametrize(
    "observable_category",
    [
        "resource_value",
        "owned_generator_count",
        "purchased_upgrade_set",
        "prestige_count",
    ],
)
def test_derive_status_runs_real_probe_classifier_for_each_observable_category(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    observable_category: str,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    simulation = _classifier_simulation(observable_category)

    def run_scenario(_simulator: object, scenario_id: str) -> SimulationResult:
        assert scenario_id == "smoke"
        return simulation

    monkeypatch.setattr(probe_module.Simulator, "run_scenario", run_scenario)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "runnable"
    assert status.structural_valid and status.smoke_eligible
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.missing_requirements == ()
    assert _manifest(project.root) == before


def test_elapsed_time_and_unlock_events_without_observable_state_are_incomplete(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    simulation = _classifier_simulation(None)

    def run_scenario(_simulator: object, scenario_id: str) -> SimulationResult:
        assert scenario_id == "smoke"
        return simulation

    monkeypatch.setattr(probe_module.Simulator, "run_scenario", run_scenario)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "incomplete"
    assert status.structural_valid and status.smoke_eligible
    assert [item.code for item in status.missing_requirements] == [
        "smoke_no_state_change"
    ]
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


@pytest.mark.parametrize(
    ("efficiencies", "weights", "expected_code"),
    [
        ({"active": "1"}, {"gather": "0"}, "activity_weight_nonpositive"),
        ({"active": "1"}, {}, "activity_weight_nonpositive"),
        ({"active": "0"}, {"gather": "1"}, "activity_efficiency_nonpositive"),
        ({"generator": "1"}, {"gather": "1"}, "activity_efficiency_nonpositive"),
    ],
)
def test_activity_route_requires_weight_and_efficiency_for_every_smoke_profile(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    efficiencies: dict[str, str],
    weights: dict[str, str],
    expected_code: str,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    _upsert(
        project,
        "player_profile",
        "second",
        {
            "source_efficiency": efficiencies,
            "behavior_policy": "cheap_unlock_first",
            "session_pattern": "authoring_default",
            "prestige_policy": "conservative",
            "activity_weights": weights,
            "luck": "1",
        },
    )
    _upsert(
        project,
        "scenario",
        "smoke",
        {
            "duration_hours": "0.002777777777777778",
            "time_mode": "tick",
            "profiles": ["default", "second"],
            "start_state": "new_player",
            "record_interval_seconds": 1,
            "outputs": ["resource_curve"],
        },
    )

    def forbidden_probe(_model: object) -> TenTickProbeResult:
        raise AssertionError("static ineligibility must not run the probe")

    monkeypatch.setattr(status_module, "run_ten_tick_probe", forbidden_probe)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "incomplete"
    assert status.structural_valid and not status.smoke_eligible
    assert expected_code in {item.code for item in status.missing_requirements}
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


def test_activity_route_with_two_positive_profiles_runs_real_probe(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    _upsert(
        project,
        "player_profile",
        "second",
        {
            "source_efficiency": {"active": "0.5"},
            "behavior_policy": "cheap_unlock_first",
            "session_pattern": "authoring_default",
            "prestige_policy": "conservative",
            "activity_weights": {"gather": "2"},
            "luck": "1",
        },
    )
    _upsert(
        project,
        "scenario",
        "smoke",
        {
            "duration_hours": "0.002777777777777778",
            "time_mode": "tick",
            "profiles": ["default", "second"],
            "start_state": "new_player",
            "record_interval_seconds": 1,
            "outputs": ["resource_curve"],
        },
    )
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "runnable"
    assert status.structural_valid and status.smoke_eligible
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


def test_latest_smoke_is_called_once_and_lookup_failures_are_status_failures(
    tmp_path: Path,
) -> None:
    project = _blank_project(tmp_path)
    calls = 0

    def broken_lookup() -> object:
        nonlocal calls
        calls += 1
        raise RuntimeError("secret backend text")

    status = derive_status(project, broken_lookup)

    assert calls == 1
    assert status.state == "failed"
    assert not status.structural_valid
    assert status.latest_smoke_run_id is None
    assert any(item.code == "latest_smoke_failed" for item in status.missing_requirements)
    assert "secret backend text" not in json.dumps(status.to_payload())


def test_latest_smoke_mapping_access_failure_is_contained(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)

    class BrokenRecord(Mapping[str, object]):
        def __getitem__(self, key: str) -> object:
            raise RuntimeError(f"secret mapping failure for {key}")

        def __iter__(self) -> Iterator[str]:
            return iter(("run_id",))

        def __len__(self) -> int:
            return 1

    status = derive_status(project, lambda: BrokenRecord())

    assert status.state == "failed"
    assert [item.code for item in status.missing_requirements] == [
        "latest_smoke_invalid"
    ]
    assert "secret mapping failure" not in json.dumps(status.to_payload())


def test_duplicate_workbook_ids_keep_partial_counts_and_discovery(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_resource(project)
    path = project.datas / "resources.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.append([None, "gold", "Gold duplicate", "currency"])
    workbook.save(path)
    workbook.close()
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 2
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(
        item.code == "duplicate_entity_id"
        and item.entity == "resource"
        and item.id == "gold"
        for item in status.missing_requirements
    )
    assert _manifest(project.root) == before


def test_duplicate_yaml_ids_keep_scenario_inventory_and_latest_smoke(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    original = project.config.read_text(encoding="utf-8")
    project.config.write_text(
        original
        + "\n  smoke:\n"
        + "    duration_hours: '1'\n"
        + "    time_mode: tick\n"
        + "    profiles: [default]\n"
        + "    start_state: new_player\n"
        + "    record_interval_seconds: 1\n"
        + "    outputs: []\n",
        encoding="utf-8",
    )

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["scenario"] == 2
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(item.code == "invalid_yaml_source" for item in status.missing_requirements)
    assert any(item.code == "duplicate_entity_id" for item in status.missing_requirements)


def test_unsafe_yaml_formula_fails_without_losing_other_counts(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    project.config.write_text(
        project.config.read_text(encoding="utf-8").replace(
            "base_cost * pow(growth, owned)",
            "__import__('os')",
        ),
        encoding="utf-8",
    )

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["formula"] == 3
    assert status.entity_counts["scenario"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(item.code == "invalid_change" for item in status.missing_requirements)


def test_unsafe_workbook_formula_keeps_readable_row_count(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_resource(project)
    path = project.datas / "resources.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["C4"] = "=1+1"
    workbook.save(path)
    workbook.close()

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(item.code == "invalid_workbook_source" for item in status.missing_requirements)


def test_malformed_yaml_after_scenarios_preserves_discoverable_ids(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    project.config.write_text(
        project.config.read_text(encoding="utf-8") + "\nbroken: [\n",
        encoding="utf-8",
    )

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.available_scenarios == ("smoke",)
    assert status.entity_counts["scenario"] == 1
    assert status.latest_smoke_run_id == "prior-smoke-1"


def test_malformed_registry_falls_back_to_canonical_partial_counts(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_resource(project)
    registry = project.datas / "__tables__.xlsx"
    workbook = load_workbook(registry)
    workbook.active["A1"] = "bad-marker"
    workbook.save(registry)
    workbook.close()

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"


def test_unsafe_registry_path_is_never_read_for_partial_counts(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    outside = project.root / "outside"
    outside.mkdir()
    outside_table = outside / "resources.xlsx"
    outside_table.write_bytes((project.datas / "resources.xlsx").read_bytes())
    workbook = load_workbook(outside_table)
    workbook.active.append([None, "outside", "Outside", "currency"])
    workbook.save(outside_table)
    workbook.close()
    registry = project.datas / "__tables__.xlsx"
    workbook = load_workbook(registry)
    workbook.active["C4"] = "../outside/resources.xlsx"
    workbook.save(registry)
    workbook.close()

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 0
    assert status.latest_smoke_run_id == "prior-smoke-1"


def test_invalid_registry_canonical_file_symlink_is_never_followed(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _break_registry(project)
    canonical = project.datas / "resources.xlsx"
    outside = tmp_path / "outside_resources.xlsx"
    _write_sentinel_resource_workbook(canonical, outside)
    canonical.unlink()
    try:
        os.symlink(outside, canonical)
    except (NotImplementedError, OSError) as error:
        pytest.skip(f"file symlink unavailable on this platform: {type(error).__name__}")
    before = _manifest(project.root)
    outside_before = outside.read_bytes()

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 0
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(
        item.code == "invalid_workbook_source" and "indirection" in item.message
        for item in status.missing_requirements
    )
    assert outside.read_bytes() == outside_before
    assert _manifest(project.root) == before


@pytest.mark.skipif(os.name != "nt", reason="Windows directory junction coverage")
def test_invalid_registry_canonical_junction_is_never_traversed(tmp_path: Path) -> None:
    is_junction = getattr(Path, "is_junction", None)
    if not callable(is_junction):
        pytest.skip("Path.is_junction is unavailable")
    project = _blank_project(tmp_path)
    _break_registry(project)
    canonical = project.datas / "resources.xlsx"
    canonical.unlink()
    outside = tmp_path / "outside_junction_target"
    outside.mkdir()
    sentinel = outside / "resources.xlsx"
    _write_sentinel_resource_workbook(project.datas / "generators.xlsx", sentinel)
    completed = subprocess.run(
        [
            os.environ.get("COMSPEC", "cmd.exe"),
            "/c",
            "mklink",
            "/J",
            str(canonical),
            str(outside),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    if completed.returncode != 0 or not canonical.is_junction():
        if canonical.exists():
            os.rmdir(canonical)
        pytest.skip("directory junction creation is unavailable")
    outside_before = sentinel.read_bytes()
    try:
        before = _manifest(project.root)

        status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

        assert status.state == "failed"
        assert status.entity_counts["resource"] == 0
        assert status.available_scenarios == ("smoke",)
        assert status.latest_smoke_run_id == "prior-smoke-1"
        assert any(
            item.code == "invalid_workbook_source" and "indirection" in item.message
            for item in status.missing_requirements
        )
        assert sentinel.read_bytes() == outside_before
        assert _manifest(project.root) == before
    finally:
        os.rmdir(canonical)


def test_canonical_fallback_retarget_before_open_is_rejected_without_reading(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = _blank_project(tmp_path)
    _break_registry(project)
    canonical = project.datas / "resources.xlsx"
    outside = tmp_path / "outside_resources.xlsx"
    _write_sentinel_resource_workbook(canonical, outside)
    original_open = status_module._open_workbook_binary
    read_attempted = False
    retargeted = False

    class TrackedHandle:
        def __init__(self, handle: object) -> None:
            self._handle = handle

        def fileno(self) -> int:
            return self._handle.fileno()  # type: ignore[attr-defined]

        def read(self, size: int = -1) -> bytes:
            nonlocal read_attempted
            read_attempted = True
            return self._handle.read(size)  # type: ignore[attr-defined]

    @contextmanager
    def retarget_before_open(selected: Path):
        nonlocal retargeted
        if selected != canonical or retargeted:
            with original_open(selected) as handle:
                yield handle
            return
        retargeted = True
        backup = canonical.with_name(".resources-original.xlsx")
        os.replace(canonical, backup)
        os.replace(outside, canonical)
        try:
            with original_open(canonical) as handle:
                yield TrackedHandle(handle)
        finally:
            os.replace(canonical, outside)
            os.replace(backup, canonical)

    monkeypatch.setattr(status_module, "_open_workbook_binary", retarget_before_open)
    before = _manifest(project.root)
    outside_before = outside.read_bytes()

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert retargeted
    assert not read_attempted
    assert status.state == "failed"
    assert status.entity_counts["resource"] == 0
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(item.code == "invalid_workbook_source" for item in status.missing_requirements)
    assert outside.read_bytes() == outside_before
    assert _manifest(project.root) == before


def test_unresolved_workbook_reference_fails_with_partial_inventory(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_resource(project)
    _upsert(
        project,
        "activity_output",
        "ghost_output",
        {
            "activity_id": "ghost",
            "output_resource": "gold",
            "amount_per_second": "1",
        },
    )

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert status.entity_counts["resource"] == 1
    assert status.entity_counts["activity_output"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert any(item.code == "status_lint_failed" for item in status.missing_requirements)


def test_ephemeral_export_is_opened_exactly_once(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    project = _blank_project(tmp_path)
    original = status_module.ephemeral_export
    calls = 0

    def counted(selected: AuthoringProject):
        nonlocal calls
        calls += 1
        return original(selected)

    monkeypatch.setattr(status_module, "ephemeral_export", counted)

    derive_status(project, lambda: None)

    assert calls == 1


@pytest.mark.parametrize(("formal", "would_be"), [(False, "runnable"), (True, "ready")])
def test_ephemeral_export_cleanup_failure_discards_computed_model_outcome(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    formal: bool,
    would_be: str,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    if formal:
        _add_formal_scenario(project)
    original = status_module.ephemeral_export

    @contextmanager
    def cleanup_failure(selected: AuthoringProject):
        with original(selected) as exported:
            yield exported
        raise RuntimeError(f"cleanup failed after {would_be}")

    monkeypatch.setattr(status_module, "ephemeral_export", cleanup_failure)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert not status.structural_valid and not status.smoke_eligible
    assert [item.code for item in status.missing_requirements] == [
        "status_export_failed"
    ]
    assert status.warnings == ()
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == (
        ("formal", "smoke") if formal else ("smoke",)
    )
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


def test_synchronized_committed_exports_have_no_stale_warning(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    with status_module.ephemeral_export(project) as current:
        shutil.copytree(current.export_root, project.exports, dirs_exist_ok=True)
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "runnable"
    assert status.warnings == ()
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


def test_corrupt_committed_export_adds_one_warning_without_changing_state(
    tmp_path: Path,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    with status_module.ephemeral_export(project) as current:
        shutil.copytree(current.export_root, project.exports, dirs_exist_ok=True)
    synchronized = derive_status(project, lambda: {"run_id": "prior-smoke-1"})
    (project.exports / "resources.json").write_bytes(b"not-json")
    before = _manifest(project.root)

    corrupt = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert corrupt.state == synchronized.state == "runnable"
    assert corrupt.structural_valid and corrupt.smoke_eligible
    assert [item.code for item in corrupt.warnings] == ["exports_stale"]
    assert corrupt.entity_counts == synchronized.entity_counts
    assert corrupt.available_scenarios == synchronized.available_scenarios == ("smoke",)
    assert corrupt.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


def test_programmer_base_exception_from_latest_lookup_is_not_swallowed(tmp_path: Path) -> None:
    project = _blank_project(tmp_path)

    with pytest.raises(KeyboardInterrupt):
        derive_status(project, lambda: (_ for _ in ()).throw(KeyboardInterrupt()))


@pytest.mark.parametrize("phase", ["build", "execution", "artifact"])
def test_each_smoke_failure_phase_is_a_structured_failed_requirement(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    phase: str,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    monkeypatch.setattr(
        status_module,
        "run_ten_tick_probe",
        lambda _model: (_ for _ in ()).throw(
            AuthoringError(
                "smoke_failed",
                f"The ten-tick smoke probe failed during {phase}.",
                {"phase": phase},
            )
        ),
    )
    before = _manifest(project.root)

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert not status.structural_valid and not status.smoke_eligible
    assert [item.code for item in status.missing_requirements] == ["smoke_failed"]
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert _manifest(project.root) == before


@pytest.mark.parametrize("phase", ["export", "load", "lint", "build", "probe"])
def test_injected_pipeline_failures_return_structured_failed_status(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    phase: str,
) -> None:
    project = _blank_project(tmp_path)
    _add_activity_route(project)
    before = _manifest(project.root)

    if phase == "export":
        def fail_export(_project: object):
            raise RuntimeError("raw export failure")
        monkeypatch.setattr(status_module, "ephemeral_export", fail_export)
    elif phase == "load":
        monkeypatch.setattr(
            status_module.ConfigLoader,
            "load",
            classmethod(lambda cls, *_args: (_ for _ in ()).throw(RuntimeError("raw load"))),
        )
    elif phase == "lint":
        monkeypatch.setattr(
            status_module.ConfigLinter,
            "validate",
            classmethod(lambda cls, *_args: (_ for _ in ()).throw(RuntimeError("raw lint"))),
        )
    elif phase == "build":
        monkeypatch.setattr(
            status_module.ModelBuilder,
            "build",
            classmethod(lambda cls, *_args: (_ for _ in ()).throw(RuntimeError("raw build"))),
        )
    else:
        monkeypatch.setattr(
            status_module,
            "run_ten_tick_probe",
            lambda _model: (_ for _ in ()).throw(
                AuthoringError(
                    "smoke_failed",
                    "The ten-tick smoke probe failed during artifact.",
                    {"phase": "artifact"},
                )
            ),
        )

    status = derive_status(project, lambda: {"run_id": "prior-smoke-1"})

    assert status.state == "failed"
    assert not status.structural_valid and not status.smoke_eligible
    assert status.entity_counts["resource"] == 1
    assert status.available_scenarios == ("smoke",)
    assert status.latest_smoke_run_id == "prior-smoke-1"
    assert status.missing_requirements
    assert _manifest(project.root) == before


def test_issue_order_is_exact_and_semantic_duplicates_are_removed() -> None:
    items = [
        EligibilityFinding("z", "m", "resource", "b"),
        EligibilityFinding("a", "z"),
        EligibilityFinding("z", "a", "resource", "a"),
        EligibilityFinding("z", "m", "resource", "b"),
    ]

    ordered = status_module._ordered_issues(items)

    assert [(item.code, item.entity, item.id, item.message) for item in ordered] == [
        ("a", None, None, "z"),
        ("z", "resource", "a", "a"),
        ("z", "resource", "b", "m"),
    ]
