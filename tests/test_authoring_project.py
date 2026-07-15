from __future__ import annotations

from dataclasses import FrozenInstanceError
import hashlib
import os
from pathlib import Path
import re
import shutil
import subprocess

from openpyxl import Workbook, load_workbook
import pytest
import yaml

from igess.authoring import AuthoringProject
from igess.authoring import project as project_module
from igess.authoring import templates as templates_module
from igess.authoring.response import AuthoringError
from igess.authoring.templates import (
    _model_id_from_output_name,
    initialize_authoring_project,
)


def _write_registry(datas: Path, paths: list[object]) -> Path:
    datas.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["##var", "table", "path", "mode", "key"])
    sheet.append(["##", None, None, None, None])
    sheet.append(["##type", "string", "string", "string", "string"])
    for index, path in enumerate(paths):
        sheet.append([None, f"table_{index}", path, "map", "id"])
    registry = datas / "__tables__.xlsx"
    workbook.save(registry)
    return registry


def _write_registry_rows(datas: Path, headers: list[object], rows: list[list[object]]) -> Path:
    datas.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["##var", *headers])
    sheet.append(["##", *([None] * len(headers))])
    sheet.append(["##type", *(["string"] * len(headers))])
    for row in rows:
        sheet.append([None, *row])
    registry = datas / "__tables__.xlsx"
    workbook.save(registry)
    return registry


def _symlink_or_skip(link: Path, target: Path, *, is_directory: bool) -> None:
    try:
        link.symlink_to(target, target_is_directory=is_directory)
    except OSError as error:
        pytest.skip(f"filesystem symlinks are unavailable: {error}")


def _make_project(root: Path, paths: list[str] | None = None) -> AuthoringProject:
    root.mkdir(parents=True, exist_ok=True)
    (root / "economy.yaml").write_text("version: 1\n", encoding="utf-8")
    registrations = paths if paths is not None else ["resources.xlsx"]
    _write_registry(root / "Datas", registrations)
    for path in registrations:
        workbook_path = root / "Datas" / path
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook_path.write_bytes(f"workbook:{path}".encode("utf-8"))
    (root / "luban_exports").mkdir()
    return AuthoringProject.discover(root)


def _expected_digest(root: Path, paths: list[Path]) -> str:
    digest = hashlib.sha256()
    for path in sorted(paths, key=lambda item: item.relative_to(root).as_posix()):
        relative = path.relative_to(root).as_posix()
        digest.update(relative.encode("utf-8"))
        digest.update(b"\0")
        digest.update(path.read_bytes())
    return f"sha256:{digest.hexdigest()}"


_BLANK_TABLE_SCHEMAS = {
    "resources.xlsx": (
        ("id", "name", "dimension"),
        ("stable resource id", "display name", "quantity dimension"),
        ("string", "string", "string"),
    ),
    "generators.xlsx": (
        (
            "id",
            "name",
            "generator_type",
            "output_resource",
            "source_type",
            "base_output",
            "base_cost",
            "cost_resource",
            "cost_growth",
            "unlock_condition",
        ),
        (
            "stable generator id",
            "display name",
            "YAML generator type",
            "produced resource id",
            "source type id",
            "base output per second",
            "first purchase cost",
            "resource spent",
            "exponential cost growth",
            "deterministic unlock condition",
        ),
        ("string",) * 10,
    ),
    "activities.xlsx": (
        ("id", "name", "source_type", "unlock_condition"),
        (
            "stable activity id",
            "display name",
            "source type id",
            "deterministic unlock condition",
        ),
        ("string",) * 4,
    ),
    "activity_outputs.xlsx": (
        ("id", "activity_id", "output_resource", "amount_per_second"),
        (
            "stable output id",
            "activity id",
            "produced resource id",
            "full-time amount per second",
        ),
        ("string",) * 4,
    ),
    "upgrades.xlsx": (
        (
            "id",
            "name",
            "target",
            "modifier_type",
            "value",
            "cost_resource",
            "base_cost",
            "unlock_condition",
        ),
        (
            "stable upgrade id",
            "display name",
            "modifier target",
            "modifier type id",
            "modifier value",
            "resource spent",
            "purchase cost",
            "deterministic unlock condition",
        ),
        ("string",) * 8,
    ),
    "constants.xlsx": (
        ("id", "value"),
        ("stable constant id", "string-encoded number"),
        ("string", "string"),
    ),
    "milestones.xlsx": (
        ("id", "name", "condition", "reward_resource", "reward_amount"),
        (
            "stable milestone id",
            "display name",
            "condition",
            "resource rewarded",
            "reward amount",
        ),
        ("string",) * 5,
    ),
    "prestige_layers.xlsx": (
        (
            "id",
            "name",
            "trigger_resource",
            "reward_resource",
            "formula",
            "divisor",
            "exponent",
            "min_gain",
            "reset_resources",
            "unlock_condition",
        ),
        (
            "stable prestige id",
            "display name",
            "resource measured",
            "resource rewarded",
            "YAML formula id",
            "formula divisor",
            "formula exponent",
            "minimum gain",
            "resources reset",
            "condition",
        ),
        (
            "string",
            "string",
            "string",
            "string",
            "string",
            "string",
            "string",
            "string",
            "(list#sep=;),string",
            "string",
        ),
    ),
}

_BLANK_REGISTRATIONS = (
    ("resources", "resources.xlsx", "map", "id"),
    ("generators", "generators.xlsx", "map", "id"),
    ("activities", "activities.xlsx", "map", "id"),
    ("activity_outputs", "activity_outputs.xlsx", "map", "id"),
    ("upgrades", "upgrades.xlsx", "map", "id"),
    ("constants", "constants.xlsx", "map", "id"),
    ("milestones", "milestones.xlsx", "map", "id"),
    ("prestige_layers", "prestige_layers.xlsx", "map", "id"),
)


def test_discover_returns_frozen_canonical_direct_child_paths(tmp_path: Path) -> None:
    requested = tmp_path / "wrapper" / ".." / "model"
    project = _make_project(tmp_path / "model")
    discovered = AuthoringProject.discover(requested)
    root = (tmp_path / "model").resolve()

    assert discovered == project
    assert discovered.root == root
    assert discovered.config == root / "economy.yaml"
    assert discovered.datas == root / "Datas"
    assert discovered.exports == root / "luban_exports"
    assert discovered.runs == root / "runs"
    assert discovered.legacy_runs == root / ".igess" / "runs"
    assert discovered.reports == root / "reports"
    assert discovered.changes == root / "changes"
    assert discovered.transactions == root / ".igess" / "transactions"
    assert discovered.lock == root / ".igess" / "model.lock"
    with pytest.raises(FrozenInstanceError):
        discovered.root = tmp_path  # type: ignore[misc]


def test_direct_construction_canonicalizes_root_and_derives_every_path(tmp_path: Path) -> None:
    discovered = _make_project(tmp_path / "model")

    constructed = AuthoringProject(tmp_path / "wrapper" / ".." / "model")

    assert constructed == discovered
    assert all(
        isinstance(getattr(constructed, field), Path)
        for field in (
            "root",
            "config",
            "datas",
            "exports",
            "runs",
            "legacy_runs",
            "reports",
            "changes",
            "transactions",
            "lock",
        )
    )


def test_direct_construction_rejects_invalid_root_without_leaking_attribute_errors() -> None:
    with pytest.raises(TypeError, match="root"):
        AuthoringProject(object())  # type: ignore[arg-type]


def test_discovery_does_not_search_descendants(tmp_path: Path) -> None:
    _make_project(tmp_path / "outer" / "nested")

    with pytest.raises(AuthoringError) as captured:
        AuthoringProject.discover(tmp_path / "outer")

    assert captured.value.code == "project_config_missing"
    assert captured.value.details == {
        "expected": "file",
        "path": str((tmp_path / "outer" / "economy.yaml").resolve()),
        "reason": "missing",
        "role": "project config",
    }


@pytest.mark.parametrize(
    ("name", "is_directory", "code", "role"),
    [
        ("economy.yaml", False, "project_config_unsafe", "project config"),
        ("Datas", True, "source_tables_unsafe", "source tables"),
        ("luban_exports", True, "runtime_exports_unsafe", "runtime exports"),
    ],
)
def test_discovery_rejects_required_child_symlinks_outside_root(
    tmp_path: Path,
    name: str,
    is_directory: bool,
    code: str,
    role: str,
) -> None:
    root = tmp_path / "model"
    _make_project(root)
    candidate = root / name
    if candidate.is_dir():
        shutil.rmtree(candidate)
    else:
        candidate.unlink()
    outside = tmp_path / f"outside-{name.replace('.', '-')}"
    if is_directory:
        outside.mkdir()
    else:
        outside.write_text("version: 1\n", encoding="utf-8")
    _symlink_or_skip(candidate, outside, is_directory=is_directory)

    with pytest.raises(AuthoringError) as captured:
        AuthoringProject.discover(root)

    assert captured.value.code == code
    assert captured.value.details["reason"] == "outside_root"
    assert captured.value.details["role"] == role
    assert captured.value.details["path"] == str(candidate)
    assert captured.value.details["resolved_path"] == str(outside.resolve())


def test_discovery_rejects_required_child_target_that_is_not_direct(tmp_path: Path) -> None:
    root = tmp_path / "model"
    _make_project(root)
    project_config = root / "economy.yaml"
    project_config.unlink()
    nested_config = root / "nested" / "economy.yaml"
    nested_config.parent.mkdir()
    nested_config.write_text("version: 1\n", encoding="utf-8")
    _symlink_or_skip(project_config, nested_config, is_directory=False)

    with pytest.raises(AuthoringError) as captured:
        AuthoringProject.discover(root)

    assert captured.value.code == "project_config_unsafe"
    assert captured.value.details["reason"] == "not_direct_child"


def test_model_digest_rejects_config_retargeted_outside_root(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model")
    outside = tmp_path / "outside.yaml"
    outside.write_text("secret: true\n", encoding="utf-8")
    project.config.unlink()
    _symlink_or_skip(project.config, outside, is_directory=False)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert captured.value.code == "project_config_unsafe"
    assert captured.value.details["reason"] == "outside_root"


@pytest.mark.parametrize(
    ("name", "expected", "role", "missing_code", "wrong_type_code"),
    [
        ("economy.yaml", "file", "project config", "project_config_missing", "project_config_wrong_type"),
        ("Datas", "directory", "source tables", "source_tables_missing", "source_tables_wrong_type"),
        (
            "luban_exports",
            "directory",
            "runtime exports",
            "runtime_exports_missing",
            "runtime_exports_wrong_type",
        ),
    ],
)
def test_discovery_reports_each_missing_and_wrong_type_role(
    tmp_path: Path,
    name: str,
    expected: str,
    role: str,
    missing_code: str,
    wrong_type_code: str,
) -> None:
    root = tmp_path / name.replace(".", "-")
    _make_project(root)
    candidate = root / name
    if candidate.is_dir():
        shutil.rmtree(candidate)
    else:
        candidate.unlink()

    with pytest.raises(AuthoringError) as missing:
        AuthoringProject.discover(root)
    assert missing.value.code == missing_code
    assert missing.value.details == {
        "expected": expected,
        "path": str(candidate.resolve()),
        "reason": "missing",
        "role": role,
    }

    if expected == "file":
        candidate.mkdir()
    else:
        candidate.write_text("not a directory", encoding="utf-8")
    with pytest.raises(AuthoringError) as wrong_type:
        AuthoringProject.discover(root)
    assert wrong_type.value.code == wrong_type_code
    assert wrong_type.value.details["role"] == role
    assert wrong_type.value.details["reason"] == "wrong_type"
    assert wrong_type.value.details["expected"] == expected
    assert wrong_type.value.details["path"] == str(candidate.resolve())


def test_read_run_roots_returns_only_existing_directories_in_modern_first_order(
    tmp_path: Path,
) -> None:
    project = _make_project(tmp_path / "model")
    assert project.read_run_roots() == []

    project.legacy_runs.mkdir(parents=True)
    assert project.read_run_roots() == [project.legacy_runs]

    project.runs.mkdir()
    assert project.read_run_roots() == [project.runs, project.legacy_runs]

    project.legacy_runs.rmdir()
    project.legacy_runs.write_text("not a directory", encoding="utf-8")
    assert project.read_run_roots() == [project.runs]


def test_read_run_roots_deduplicates_directory_aliases(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model")
    project.runs.mkdir()
    project.legacy_runs.parent.mkdir()
    try:
        project.legacy_runs.symlink_to(project.runs, target_is_directory=True)
    except OSError as error:
        pytest.skip(f"directory symlinks are unavailable: {error}")

    assert project.read_run_roots() == [project.runs]


def test_read_run_roots_skips_paths_that_cannot_be_resolved(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model")
    project.runs.mkdir()
    project.legacy_runs.mkdir(parents=True)
    original_resolve = Path.resolve

    def flaky_resolve(path: Path, *args: object, **kwargs: object) -> Path:
        if path == project.runs:
            raise OSError("retargeted while reading")
        return original_resolve(path, *args, **kwargs)

    monkeypatch.setattr(Path, "resolve", flaky_resolve)

    assert project.read_run_roots() == [project.legacy_runs]


def test_model_digest_hashes_canonical_sorted_registered_sources_exactly(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model", ["zeta.xlsx", "nested/alpha.xlsx"])
    expected_paths = [
        project.config,
        project.datas / "__tables__.xlsx",
        project.datas / "zeta.xlsx",
        project.datas / "nested" / "alpha.xlsx",
    ]

    value = project.model_digest()

    assert value == _expected_digest(project.root, expected_paths)
    assert re.fullmatch(r"sha256:[0-9a-f]{64}", value)
    assert project.model_digest() == value


def test_model_digest_streams_files_with_fixed_bounded_reads(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model")
    project.config.write_bytes(b"x" * (3 * 1024 * 1024 + 17))
    expected = project.model_digest()
    original_open = Path.open
    read_sizes: list[int] = []

    class TrackingReader:
        def __init__(self, stream: object) -> None:
            self._stream = stream

        def __enter__(self) -> TrackingReader:
            return self

        def __exit__(self, *args: object) -> None:
            self.close()

        def fileno(self) -> int:
            return self._stream.fileno()  # type: ignore[attr-defined,no-any-return]

        def seek(self, offset: int, whence: int = 0) -> int:
            return self._stream.seek(offset, whence)  # type: ignore[attr-defined,no-any-return]

        def close(self) -> None:
            self._stream.close()  # type: ignore[attr-defined]

        def read(self, size: int = -1) -> bytes:
            read_sizes.append(size)
            assert size > 0
            return self._stream.read(size)  # type: ignore[attr-defined,no-any-return]

    def tracking_open(path: Path, *args: object, **kwargs: object) -> object:
        if path == project.config:
            return TrackingReader(original_open(path, *args, **kwargs))
        return original_open(path, *args, **kwargs)

    monkeypatch.setattr(Path, "open", tracking_open)

    assert project.model_digest() == expected
    assert len(read_sizes) > 2
    assert len(set(read_sizes)) == 1


def test_model_digest_converts_memory_error_to_structured_diagnostic(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model")
    original_open = Path.open

    class ExhaustedReader:
        def __init__(self, stream: object) -> None:
            self._stream = stream

        def __enter__(self) -> ExhaustedReader:
            return self

        def __exit__(self, *args: object) -> None:
            self.close()

        def fileno(self) -> int:
            return self._stream.fileno()  # type: ignore[attr-defined,no-any-return]

        def seek(self, offset: int, whence: int = 0) -> int:
            return self._stream.seek(offset, whence)  # type: ignore[attr-defined,no-any-return]

        def close(self) -> None:
            self._stream.close()  # type: ignore[attr-defined]

        def read(self, size: int = -1) -> bytes:
            raise MemoryError("simulated allocation failure")

    def exhausted_open(path: Path, *args: object, **kwargs: object) -> object:
        if path == project.config:
            return ExhaustedReader(original_open(path, *args, **kwargs))
        return original_open(path, *args, **kwargs)

    monkeypatch.setattr(Path, "open", exhausted_open)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert captured.value.code == "project_config_unreadable"
    assert captured.value.details["reason"] == "source_read_error"
    assert captured.value.details["error_type"] == "MemoryError"


def test_model_digest_rejects_config_identity_swap_before_open(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model")
    replacement = tmp_path / "replacement-config.yaml"
    replacement.write_text("external: true\n", encoding="utf-8")
    original_open = getattr(project_module, "_open_binary", lambda path: path.open("rb"))
    swapped = False

    def swapping_open(path: Path) -> object:
        nonlocal swapped
        if path == project.config and not swapped:
            swapped = True
            os.replace(replacement, path)
        return original_open(path)

    monkeypatch.setattr(project_module, "_open_binary", swapping_open, raising=False)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert swapped
    assert captured.value.code == "project_config_unsafe"
    assert captured.value.details["reason"] == "path_identity_changed"


def test_model_digest_rejects_registry_identity_swap_before_parsing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model")
    replacement_root = tmp_path / "replacement-registry"
    replacement = _write_registry(replacement_root, ["external.xlsx"])
    original_open = getattr(project_module, "_open_binary", lambda path: path.open("rb"))
    registry = project.datas / "__tables__.xlsx"
    swapped = False

    def swapping_open(path: Path) -> object:
        nonlocal swapped
        if path == registry and not swapped:
            swapped = True
            os.replace(replacement, path)
        return original_open(path)

    monkeypatch.setattr(project_module, "_open_binary", swapping_open, raising=False)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert swapped
    assert captured.value.code == "invalid_source_registry"
    assert captured.value.details["reason"] == "path_identity_changed"


def test_model_digest_rejects_registered_source_identity_swap_before_consumption(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    project = _make_project(tmp_path / "model", ["source.xlsx"])
    replacement = tmp_path / "replacement-source.xlsx"
    replacement.write_bytes(b"external source bytes")
    original_open = getattr(project_module, "_open_binary", lambda path: path.open("rb"))
    source = project.datas / "source.xlsx"
    swapped = False

    def swapping_open(path: Path) -> object:
        nonlocal swapped
        if path == source and not swapped:
            swapped = True
            os.replace(replacement, path)
        return original_open(path)

    monkeypatch.setattr(project_module, "_open_binary", swapping_open, raising=False)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert swapped
    assert captured.value.code == "invalid_source_registry"
    assert captured.value.details["reason"] == "path_identity_changed"


def test_model_digest_changes_for_config_registry_and_registered_workbook(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model", ["resources.xlsx"])
    original = project.model_digest()

    project.config.write_text("version: 2\n", encoding="utf-8")
    config_changed = project.model_digest()
    assert config_changed != original

    project.datas.joinpath("resources.xlsx").write_bytes(b"changed source")
    source_changed = project.model_digest()
    assert source_changed != config_changed

    registry = load_workbook(project.datas / "__tables__.xlsx")
    registry.active["B4"] = "renamed_table"
    registry.save(project.datas / "__tables__.xlsx")
    assert project.model_digest() != source_changed


def test_model_digest_ignores_unregistered_and_derived_artifacts(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model", ["resources.xlsx"])
    original = project.model_digest()

    ignored_files = [
        project.datas / "unregistered.xlsx",
        project.exports / "resources.json",
        project.runs / "run-1" / "status.json",
        project.reports / "index.html",
        project.changes / "change-1.json",
        project.transactions / "tx-1" / "journal.json",
        project.lock,
    ]
    for index, path in enumerate(ignored_files):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(f"ignored:{index}".encode("utf-8"))

    assert project.model_digest() == original


@pytest.mark.parametrize(
    ("registrations", "reason"),
    [
        ([None], "missing_registration_path"),
        (["../outside.xlsx"], "unsafe_registration_path"),
        (["missing.xlsx"], "registered_source_missing"),
        (["one.xlsx", "./one.xlsx"], "duplicate_registration_path"),
    ],
)
def test_model_digest_rejects_invalid_registration_paths(
    tmp_path: Path,
    registrations: list[object],
    reason: str,
) -> None:
    root = tmp_path / reason
    root.mkdir()
    (root / "economy.yaml").write_text("version: 1\n", encoding="utf-8")
    _write_registry(root / "Datas", registrations)
    (root / "Datas" / "one.xlsx").write_bytes(b"one")
    (root / "luban_exports").mkdir()
    project = AuthoringProject.discover(root)

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert captured.value.code == "invalid_source_registry"
    assert captured.value.details["reason"] == reason
    assert captured.value.details["role"] == "source registry"
    assert isinstance(captured.value.details["path"], str)


def test_model_digest_rejects_absolute_duplicate_and_wrong_type_sources(tmp_path: Path) -> None:
    root = tmp_path / "model"
    project = _make_project(root, ["source.xlsx"])

    _write_registry(project.datas, [str((tmp_path / "outside.xlsx").resolve())])
    with pytest.raises(AuthoringError) as absolute:
        project.model_digest()
    assert absolute.value.details["reason"] == "unsafe_registration_path"

    _write_registry(project.datas, ["source.xlsx", "SOURCE.xlsx"])
    with pytest.raises(AuthoringError) as duplicate:
        project.model_digest()
    assert duplicate.value.details["reason"] == "duplicate_registration_path"

    _write_registry(project.datas, ["directory.xlsx"])
    (project.datas / "directory.xlsx").mkdir()
    with pytest.raises(AuthoringError) as wrong_type:
        project.model_digest()
    assert wrong_type.value.details["reason"] == "registered_source_wrong_type"


@pytest.mark.parametrize(
    ("headers", "row", "header_issue"),
    [
        (["path", "mode", "key"], ["source.xlsx", "map", "id"], "missing_required_header"),
        (
            ["table", "table", "path"],
            ["resources", "resources_copy", "source.xlsx"],
            "duplicate_header",
        ),
        (
            ["table", "path", "path"],
            ["resources", "source.xlsx", "source.xlsx"],
            "duplicate_header",
        ),
        (
            ["table", "path", "mode", "mode"],
            ["resources", "source.xlsx", "map", "map"],
            "duplicate_header",
        ),
    ],
)
def test_model_digest_rejects_missing_or_duplicate_registry_headers(
    tmp_path: Path,
    headers: list[object],
    row: list[object],
    header_issue: str,
) -> None:
    project = _make_project(tmp_path / "model", ["source.xlsx"])
    _write_registry_rows(project.datas, headers, [row])

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert captured.value.code == "invalid_source_registry"
    assert captured.value.details["reason"] == "malformed_registry"
    assert captured.value.details["header_issue"] == header_issue


def test_model_digest_rejects_hard_linked_duplicate_registrations(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model", ["source.xlsx"])
    alias = project.datas / "alias.xlsx"
    try:
        os.link(project.datas / "source.xlsx", alias)
    except OSError as error:
        pytest.skip(f"filesystem hard links are unavailable: {error}")
    _write_registry(project.datas, ["source.xlsx", "alias.xlsx"])

    with pytest.raises(AuthoringError) as captured:
        project.model_digest()

    assert captured.value.code == "invalid_source_registry"
    assert captured.value.details["reason"] == "duplicate_registration_path"


def test_model_digest_wraps_missing_or_malformed_registry_as_structured_error(tmp_path: Path) -> None:
    project = _make_project(tmp_path / "model")
    registry = project.datas / "__tables__.xlsx"
    registry.unlink()

    with pytest.raises(AuthoringError) as missing:
        project.model_digest()
    assert missing.value.code == "invalid_source_registry"
    assert missing.value.details == {
        "path": str(registry),
        "reason": "registry_missing",
        "role": "source registry",
    }

    registry.write_text("not an xlsx file", encoding="utf-8")
    with pytest.raises(AuthoringError) as malformed:
        project.model_digest()
    assert malformed.value.code == "invalid_source_registry"
    assert malformed.value.details["reason"] == "malformed_registry"
    assert isinstance(malformed.value.details["error_type"], str)
    assert all(isinstance(value, (str, int, bool)) or value is None for value in malformed.value.details.values())


def test_initialize_authoring_project_creates_exact_blank_tree_and_is_discoverable(
    tmp_path: Path,
) -> None:
    target = tmp_path / "My Game.v1!"

    created = initialize_authoring_project(target)

    assert created == target
    assert sorted(path.relative_to(target).as_posix() for path in target.iterdir()) == [
        "Datas",
        "README.md",
        "changes",
        "economy.yaml",
        "luban_exports",
        "reports",
        "run.ps1",
        "runs",
    ]
    assert sorted(path.name for path in (target / "Datas").iterdir()) == [
        "__tables__.xlsx",
        "activities.xlsx",
        "activity_outputs.xlsx",
        "constants.xlsx",
        "generators.xlsx",
        "milestones.xlsx",
        "prestige_layers.xlsx",
        "resources.xlsx",
        "upgrades.xlsx",
    ]
    for relative in ("luban_exports", "runs", "reports", "changes"):
        assert list((target / relative).iterdir()) == []
    assert AuthoringProject.discover(target).root == target.resolve()


def test_initialize_authoring_project_accepts_an_existing_empty_directory(
    tmp_path: Path,
) -> None:
    target = tmp_path / "empty"
    target.mkdir()

    assert initialize_authoring_project(target, model_id="valid-ID_9") == target
    assert yaml.safe_load((target / "economy.yaml").read_text(encoding="utf-8"))["model"][
        "id"
    ] == "valid-ID_9"
    assert list(tmp_path.iterdir()) == [target]


def test_initialize_authoring_project_restores_original_directory_if_backup_rename_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "empty"
    target.mkdir()
    os.utime(target, ns=(1_700_000_000_000_000_000,) * 2)
    original = target.stat()
    injected = OSError("injected backup rename failure")
    real_rename = os.rename

    def fail_backup_rename(source: object, destination: object) -> None:
        if Path(source) == target:
            raise injected
        real_rename(source, destination)

    monkeypatch.setattr(templates_module.os, "rename", fail_backup_rename)

    with pytest.raises(OSError) as captured:
        initialize_authoring_project(target)

    restored = target.stat()
    assert captured.value is injected
    assert os.path.samestat(original, restored)
    assert restored.st_mtime_ns == original.st_mtime_ns
    assert list(target.iterdir()) == []
    assert list(tmp_path.iterdir()) == [target]


def test_initialize_authoring_project_restores_original_directory_if_install_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "empty"
    target.mkdir()
    os.utime(target, ns=(1_700_000_000_000_000_000,) * 2)
    original = target.stat()
    injected = OSError("injected staged install failure")
    real_replace = os.replace

    def fail_staged_install(source: object, destination: object) -> None:
        if Path(destination) == target:
            raise injected
        real_replace(source, destination)

    monkeypatch.setattr(templates_module.os, "replace", fail_staged_install)

    with pytest.raises(OSError) as captured:
        initialize_authoring_project(target)

    restored = target.stat()
    assert captured.value is injected
    assert os.path.samestat(original, restored)
    assert restored.st_mtime_ns == original.st_mtime_ns
    assert list(target.iterdir()) == []
    assert list(tmp_path.iterdir()) == [target]


def test_initialize_authoring_project_preserves_raced_content_when_revalidating_backup(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "empty"
    target.mkdir()
    original = target.stat()
    real_rename = os.rename

    def add_content_after_backup_rename(source: object, destination: object) -> None:
        real_rename(source, destination)
        if Path(source) == target:
            (Path(destination) / "late.txt").write_bytes(b"do not delete")

    monkeypatch.setattr(
        templates_module.os, "rename", add_content_after_backup_rename
    )

    with pytest.raises(AuthoringError) as captured:
        initialize_authoring_project(target)

    assert captured.value.code == "project_not_empty"
    assert captured.value.details["reason"] == "not_empty"
    assert os.path.samestat(original, target.stat())
    assert (target / "late.txt").read_bytes() == b"do not delete"
    assert list(tmp_path.iterdir()) == [target]


@pytest.mark.skipif(os.name != "nt", reason="directory junctions are Windows-only")
def test_initialize_authoring_project_rejects_directory_junction_without_replacing_it(
    tmp_path: Path,
) -> None:
    actual = tmp_path / "actual"
    actual.mkdir()
    sentinel = actual / "sentinel.txt"
    sentinel.write_bytes(b"keep")
    junction = tmp_path / "junction"
    created = subprocess.run(
        ["cmd", "/c", "mklink", "/J", str(junction), str(actual)],
        check=False,
        capture_output=True,
        text=True,
    )
    if created.returncode != 0:
        pytest.skip(f"directory junctions unavailable: {created.stderr or created.stdout}")

    try:
        with pytest.raises(AuthoringError) as captured:
            initialize_authoring_project(junction)

        assert captured.value.code == "project_not_empty"
        assert captured.value.details == {
            "path": str(junction),
            "reason": "unsafe_reparse_point",
        }
        assert junction.exists()
        assert sentinel.read_bytes() == b"keep"
    finally:
        try:
            junction.lstat()
        except FileNotFoundError:
            pass
        else:
            os.rmdir(junction)


@pytest.mark.parametrize("existing_kind", ["directory", "file"])
def test_initialize_authoring_project_refuses_nonempty_target_without_changes(
    tmp_path: Path, existing_kind: str
) -> None:
    target = tmp_path / "occupied"
    if existing_kind == "directory":
        target.mkdir()
        sentinel = target / "keep.txt"
        sentinel.write_bytes(b"do not touch")
    else:
        target.write_bytes(b"do not touch")
        sentinel = target
    before = {path: path.read_bytes() for path in tmp_path.rglob("*") if path.is_file()}

    with pytest.raises(AuthoringError) as captured:
        initialize_authoring_project(target)

    assert captured.value.code == "project_not_empty"
    assert captured.value.details["path"] == str(target)
    assert captured.value.details["reason"] in {"not_empty", "not_directory"}
    assert sentinel.read_bytes() == b"do not touch"
    assert {path: path.read_bytes() for path in tmp_path.rglob("*") if path.is_file()} == before


@pytest.mark.parametrize("model_id", ["", "has space", "dot.bad", "中文", "a/b"])
def test_initialize_authoring_project_rejects_invalid_explicit_id_before_writing(
    tmp_path: Path, model_id: str
) -> None:
    target = tmp_path / "must-not-exist"

    with pytest.raises(AuthoringError) as captured:
        initialize_authoring_project(target, model_id=model_id)

    assert captured.value.code == "invalid_model_id"
    assert captured.value.details == {
        "allowed": "[A-Za-z0-9_-]+",
        "model_id": model_id,
        "path": str(target),
        "reason": "invalid_explicit_id",
    }
    assert not target.exists()


@pytest.mark.parametrize(
    ("output_name", "expected"),
    [
        ("plain-ID_9", "plain-ID_9"),
        ("My Game.v1!", "My_Game_v1_"),
        ("!!!", "___"),
        ("", "model"),
    ],
)
def test_default_model_id_sanitizes_output_name(output_name: str, expected: str) -> None:
    assert _model_id_from_output_name(output_name) == expected


def test_initialize_authoring_project_writes_exact_engine_default_yaml(
    tmp_path: Path,
) -> None:
    target = tmp_path / "blank"
    initialize_authoring_project(target)

    raw = (target / "economy.yaml").read_bytes()
    data = yaml.safe_load(raw)

    assert b"\r" not in raw
    assert raw.endswith(b"\n")
    assert data == {
        "model": {
            "id": "blank",
            "tick_seconds": 1,
            "number_backend": "bignum_log",
            "random_seed": 20260626,
        },
        "formulas": {
            "exponential_cost": {
                "args": ["base_cost", "growth", "owned"],
                "expr": "base_cost * pow(growth, owned)",
            },
            "generator_output": {
                "args": ["base_output", "owned", "multiplier"],
                "expr": "base_output * owned * multiplier",
            },
            "prestige_gain": {
                "args": ["progress", "divisor", "exponent"],
                "expr": "floor(pow(progress / divisor, exponent))",
            },
        },
        "generator_types": {
            "building": {
                "cost_formula": "exponential_cost",
                "production_formula": "generator_output",
            }
        },
        "source_types": {
            "active": {"description": "Active player actions"},
            "generator": {"description": "Automatic generator output"},
            "offline": {"description": "Offline reward"},
            "milestone": {"description": "Milestone reward"},
            "prestige": {"description": "Prestige reward"},
        },
        "modifier_pipeline": {"order": ["base", "flat", "add_pct", "mult", "exp"]},
        "modifier_types": {
            "flat": {"stage": "flat"},
            "add_pct": {"stage": "add_pct"},
            "multiply": {"stage": "mult"},
            "exponent": {"stage": "exp"},
        },
        "behavior_policies": {"cheap_unlock_first": {"type": "cheap_unlock_first"}},
        "session_patterns": {
            "authoring_default": {
                "offline_every_seconds": 60,
                "offline_duration_seconds": 0,
            }
        },
        "player_profiles": {
            "default": {
                "source_efficiency": {
                    "active": "1",
                    "generator": "1",
                    "offline": "1",
                    "milestone": "1",
                    "prestige": "1",
                },
                "behavior_policy": "cheap_unlock_first",
                "session_pattern": "authoring_default",
                "prestige_policy": "conservative",
                "activity_weights": {},
                "luck": "1",
            }
        },
        "scenarios": {
            "smoke": {
                "duration_hours": "0.002777777777777778",
                "time_mode": "tick",
                "profiles": ["default"],
                "start_state": "new_player",
                "record_interval_seconds": 1,
                "outputs": [
                    "resource_curve",
                    "purchase_timeline",
                    "unlock_timeline",
                    "prestige_timeline",
                    "bottleneck_report",
                ],
            }
        },
    }
    duration = float(data["scenarios"]["smoke"]["duration_hours"])
    assert int(duration * 3600 / data["model"]["tick_seconds"]) == 10


def test_initialize_authoring_project_writes_exact_blank_luban_workbooks(
    tmp_path: Path,
) -> None:
    target = tmp_path / "blank"
    initialize_authoring_project(target)

    for filename, (headers, comments, types) in _BLANK_TABLE_SCHEMAS.items():
        workbook = load_workbook(target / "Datas" / filename, read_only=True, data_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()
        assert rows == [
            ("##var", *headers),
            ("##", *comments),
            ("##type", *types),
        ], filename

    registry = load_workbook(
        target / "Datas" / "__tables__.xlsx", read_only=True, data_only=True
    )
    try:
        rows = list(registry.active.iter_rows(values_only=True))
    finally:
        registry.close()
    assert rows[:3] == [
        ("##var", "table", "path", "mode", "key"),
        ("##", "stable table id", "source workbook", "export mode", "map key field"),
        ("##type", "string", "string", "string", "string"),
    ]
    assert rows[3:] == [(None, *registration) for registration in _BLANK_REGISTRATIONS]


def test_blank_workbook_templates_are_byte_deterministic(tmp_path: Path) -> None:
    first = tmp_path / "first"
    second = tmp_path / "second"
    initialize_authoring_project(first, model_id="same")
    initialize_authoring_project(second, model_id="same")

    assert {
        path.name: path.read_bytes() for path in (first / "Datas").iterdir()
    } == {path.name: path.read_bytes() for path in (second / "Datas").iterdir()}


def test_initialize_authoring_project_documents_agent_workflow_and_robust_runner(
    tmp_path: Path,
) -> None:
    target = tmp_path / "blank"
    initialize_authoring_project(target)

    readme_bytes = (target / "README.md").read_bytes()
    run_bytes = (target / "run.ps1").read_bytes()
    readme = readme_bytes.decode("utf-8")
    runner = run_bytes.decode("utf-8")

    assert b"\r" not in readme_bytes
    assert b"\r" not in run_bytes
    assert "economy.yaml" in readme and "Datas/" in readme
    assert "formal sources" in readme
    assert "luban_exports/" in readme and "generated" in readme
    for command in (
        "igess model init",
        "igess model status",
        "igess model apply",
        "igess model simulate",
    ):
        assert command in readme
    for artifact in ("runs/", "reports/", "changes/"):
        assert artifact in readme
    assert "Agent" in readme and "one rule" in readme
    assert "$PSScriptRoot" in runner
    assert "igess model status --project $PSScriptRoot" in runner
    assert "igess model simulate --project $PSScriptRoot --scenario smoke" in runner
    assert runner.index("model status") < runner.index("model simulate")
