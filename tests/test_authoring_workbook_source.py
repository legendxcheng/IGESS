from __future__ import annotations

from datetime import date
import json
import os
from pathlib import Path
import random
import re
import stat
from typing import Any, NoReturn
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
import pytest

from igess.authoring.change import ModelChange
from igess.authoring.entity_schema import get_entity_schema
from igess.authoring.response import AuthoringError, CommandResponse
from igess.authoring.workbook_source import (
    TableInspection,
    WorkbookRecord,
    find_duplicate_ids,
    inspect_table,
    upsert_workbook_entity,
)


ENTITY_CASES: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {
    "resource": (
        {"name": "Gold", "dimension": "currency"},
        {"name": "Golden Coin", "dimension": "coin"},
    ),
    "generator": (
        {
            "name": "Mine",
            "generator_type": "building",
            "output_resource": "gold",
            "source_type": "generator",
            "base_output": "1",
            "base_cost": "10",
            "cost_resource": "gold",
            "cost_growth": "1.15",
            "unlock_condition": "always",
        },
        {
            "name": "Deep Mine",
            "generator_type": "building",
            "output_resource": "gold",
            "source_type": "generator",
            "base_output": "2.5",
            "base_cost": "25",
            "cost_resource": "gold",
            "cost_growth": "1.2",
            "unlock_condition": "owned(mine) >= 1",
        },
    ),
    "activity": (
        {"name": "Pan", "source_type": "active", "unlock_condition": "always"},
        {
            "name": "Pan Faster",
            "source_type": "active",
            "unlock_condition": "owned(mine) >= 1",
        },
    ),
    "activity_output": (
        {
            "activity_id": "pan",
            "output_resource": "gold",
            "amount_per_second": "0.25",
        },
        {
            "activity_id": "pan_fast",
            "output_resource": "gold",
            "amount_per_second": "1.5",
        },
    ),
    "upgrade": (
        {
            "name": "Sharper Pick",
            "target": "generator:mine.output",
            "modifier_type": "multiply",
            "value": "2",
            "cost_resource": "gold",
            "base_cost": "100",
            "unlock_condition": "always",
        },
        {
            "name": "Sharpest Pick",
            "target": "generator:*.output",
            "modifier_type": "multiply",
            "value": "3",
            "cost_resource": "gold",
            "base_cost": "250",
            "unlock_condition": "owned(mine) >= 2",
        },
    ),
    "constant": ({"value": "100"}, {"value": "125.5"}),
    "milestone": (
        {
            "name": "First Mine",
            "condition": "owned(mine) >= 1",
            "reward_resource": "gold",
            "reward_amount": "50",
        },
        {
            "name": "Two Mines",
            "condition": "owned(mine) >= 2",
            "reward_resource": "gold",
            "reward_amount": "125",
        },
    ),
    "prestige_layer": (
        {
            "name": "Renown",
            "trigger_resource": "gold",
            "reward_resource": "renown",
            "formula": "prestige_gain",
            "divisor": "1000",
            "exponent": "0.5",
            "min_gain": "1",
            "reset_resources": ["gold", "ore"],
            "unlock_condition": "always",
        },
        {
            "name": "Greater Renown",
            "trigger_resource": "gold",
            "reward_resource": "renown",
            "formula": "prestige_gain",
            "divisor": "2500",
            "exponent": "0.75",
            "min_gain": "2",
            "reset_resources": ["gold", "ore", "gem"],
            "unlock_condition": "owned(mine) >= 2",
        },
    ),
}


def _expected_type(entity: str, field: str) -> str:
    if entity == "prestige_layer" and field == "reset_resources":
        return "(list#sep=;),string"
    return "string"


def _path_for(tmp_path: Path, entity: str) -> Path:
    return tmp_path / get_entity_schema(entity).storage_name


def _encode_field(entity: str, field: str, value: Any) -> Any:
    if entity == "prestige_layer" and field == "reset_resources":
        return ";".join(value)
    return value


def _write_table(
    path: Path,
    entity: str,
    rows: list[tuple[str, dict[str, Any]]] | None = None,
    *,
    marker_row: int = 1,
    marker_column: int = 1,
    extra_header: str | None = None,
) -> None:
    schema = get_entity_schema(entity)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = schema.storage_name.removesuffix(".xlsx")
    headers = ("id", *schema.field_names)
    marker_rows = (marker_row, marker_row + 1, marker_row + 2)
    for row, marker in zip(marker_rows, ("##var", "##", "##type")):
        sheet.cell(row=row, column=marker_column, value=marker)
    for offset, header in enumerate(headers, start=1):
        column = marker_column + offset
        sheet.cell(row=marker_rows[0], column=column, value=header)
        sheet.cell(row=marker_rows[1], column=column, value=f"{header} description")
        sheet.cell(
            row=marker_rows[2],
            column=column,
            value=_expected_type(entity, header),
        )
    if extra_header:
        column = marker_column + len(headers) + 1
        sheet.cell(row=marker_rows[0], column=column, value=extra_header)
        sheet.cell(row=marker_rows[1], column=column, value="unrelated")
        sheet.cell(row=marker_rows[2], column=column, value="string")
    for row_offset, (entity_id, fields) in enumerate(rows or (), start=1):
        row = marker_rows[2] + row_offset
        sheet.cell(row=row, column=marker_column + 1, value=entity_id)
        for field_offset, field in enumerate(schema.field_names, start=2):
            sheet.cell(
                row=row,
                column=marker_column + field_offset,
                value=_encode_field(entity, field, fields[field]),
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def _change(entity: str, entity_id: str, fields: dict[str, Any]) -> ModelChange:
    return ModelChange(
        version=1,
        operation="upsert",
        entity=entity,
        id=entity_id,
        fields=fields,
    )


def _fields(record: WorkbookRecord) -> dict[str, Any]:
    return dict(record.fields)


def _inspection_fields(entity: str, fields: dict[str, Any]) -> dict[str, Any]:
    expected = dict(fields)
    if entity == "prestige_layer":
        expected["reset_resources"] = tuple(expected["reset_resources"])
    return expected


def _reversed_fields(fields: dict[str, Any]) -> dict[str, Any]:
    return dict(reversed(tuple(fields.items())))


def _shuffled_fields(fields: dict[str, Any], seed: str) -> dict[str, Any]:
    items = list(fields.items())
    random.Random(seed).shuffle(items)
    if len(items) > 1 and items == list(fields.items()):
        items = items[1:] + items[:1]
    return dict(items)


@pytest.mark.parametrize("entity", ENTITY_CASES)
def test_create_and_update_every_workbook_entity(tmp_path: Path, entity: str) -> None:
    initial, updated = ENTITY_CASES[entity]
    path = _path_for(tmp_path, entity)
    _write_table(path, entity)

    assert upsert_workbook_entity(path, _change(entity, "entry", initial)) is True
    created = inspect_table(path)
    assert isinstance(created, TableInspection)
    assert created.entity == entity
    assert len(created.records) == 1
    assert created.records[0].entity_id == "entry"
    assert _fields(created.records[0]) == _inspection_fields(entity, initial)

    assert upsert_workbook_entity(path, _change(entity, "entry", updated)) is True
    inspected = inspect_table(path)
    assert len(inspected.records) == 1
    assert inspected.records[0].row == created.records[0].row
    assert _fields(inspected.records[0]) == _inspection_fields(entity, updated)


@pytest.mark.parametrize("entity", ENTITY_CASES)
def test_field_insertion_order_does_not_affect_create_noop_or_update(
    tmp_path: Path,
    entity: str,
) -> None:
    initial, updated = ENTITY_CASES[entity]
    schema = get_entity_schema(entity)
    path = _path_for(tmp_path, entity)
    _write_table(path, entity)

    assert upsert_workbook_entity(
        path,
        _change(entity, "entry", _reversed_fields(initial)),
    ) is True
    created = inspect_table(path)
    assert tuple(name for name, _ in created.records[0].fields) == schema.field_names
    before_noop = path.read_bytes()

    assert upsert_workbook_entity(
        path,
        _change(entity, "entry", _shuffled_fields(initial, entity)),
    ) is False
    assert path.read_bytes() == before_noop

    assert upsert_workbook_entity(
        path,
        _change(entity, "entry", _reversed_fields(updated)),
    ) is True
    inspected = inspect_table(path)
    assert inspected.records[0].row == created.records[0].row
    assert _fields(inspected.records[0]) == _inspection_fields(entity, updated)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    assert tuple(
        sheet.cell(inspected.var_row, column).value
        for column in range(2, 2 + len(schema.field_names) + 1)
    ) == ("id", *schema.field_names)
    workbook.close()


def test_prestige_list_is_encoded_with_semicolons_and_decoded_as_tuple(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "prestige_layer")
    fields = ENTITY_CASES["prestige_layer"][0]
    _write_table(path, "prestige_layer")

    assert upsert_workbook_entity(
        path, _change("prestige_layer", "renown", fields)
    )

    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    headers = {sheet.cell(1, column).value: column for column in range(2, sheet.max_column + 1)}
    assert sheet.cell(4, headers["reset_resources"]).value == "gold;ore"
    workbook.close()
    assert _fields(inspect_table(path).records[0])["reset_resources"] == (
        "gold",
        "ore",
    )


def test_inspection_finds_shifted_markers_without_changing_the_file(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "resource")
    fields = ENTITY_CASES["resource"][0]
    _write_table(path, "resource", [("gold", fields)], marker_row=4, marker_column=3)
    before = path.read_bytes()

    inspected = inspect_table(path)

    assert (
        inspected.marker_column,
        inspected.var_row,
        inspected.description_row,
        inspected.type_row,
    ) == (
        3,
        4,
        5,
        6,
    )
    assert dict(inspected.columns)["id"] == 4
    assert _fields(inspected.records[0]) == fields
    assert find_duplicate_ids(path) == []
    assert path.read_bytes() == before


def test_duplicate_ids_are_reported_in_source_order_and_upsert_is_rejected(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    fields = ENTITY_CASES["resource"][0]
    _write_table(
        path,
        "resource",
        [("gold", fields), ("fish", fields), ("gold", fields), ("gold", fields)],
    )
    before = path.read_bytes()

    assert find_duplicate_ids(path) == ["gold"]
    with pytest.raises(AuthoringError) as caught:
        upsert_workbook_entity(path, _change("resource", "gold", fields))

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "duplicate_ids"
    assert caught.value.details["duplicate_ids"] == ("gold",)
    assert path.read_bytes() == before


def test_semantic_noop_does_not_save_or_add_a_row(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "constant")
    _write_table(path, "constant", [("starting_gold", {"value": 100})])
    before = path.read_bytes()
    before_stat = path.stat()

    assert upsert_workbook_entity(
        path, _change("constant", "starting_gold", {"value": "100"})
    ) is False

    assert path.read_bytes() == before
    assert path.stat().st_mtime_ns == before_stat.st_mtime_ns
    workbook = load_workbook(path)
    assert workbook.active.max_row == 4
    workbook.close()


def _decorate_row(path: Path, row: int, *, start_column: int, end_column: int) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row, column)
        cell.font = Font(name=f"Font{column}", bold=True, color="112233")
        cell.fill = PatternFill("solid", fgColor=f"{column:06X}")
        cell.border = Border(left=Side(style="thin", color="445566"))
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.protection = Protection(locked=False, hidden=True)
        cell.number_format = f"0.{column}"
    workbook.save(path)
    workbook.close()


def test_new_row_copies_each_schema_cell_style_from_nearest_data_row(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    fields = ENTITY_CASES["resource"][0]
    _write_table(path, "resource", [("old", fields)])
    _decorate_row(path, 4, start_column=2, end_column=4)

    assert upsert_workbook_entity(
        path, _change("resource", "new", ENTITY_CASES["resource"][1])
    )

    workbook = load_workbook(path)
    sheet = workbook.active
    for column in range(2, 5):
        assert sheet.cell(5, column)._style == sheet.cell(4, column)._style
    workbook.close()


def test_blank_template_copies_each_schema_cell_style_from_type_row(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    _decorate_row(path, 3, start_column=2, end_column=4)

    assert upsert_workbook_entity(
        path, _change("resource", "new", ENTITY_CASES["resource"][0])
    )

    workbook = load_workbook(path)
    sheet = workbook.active
    for column in range(2, 5):
        assert sheet.cell(4, column)._style == sheet.cell(3, column)._style
    workbook.close()


def test_update_preserves_formulas_comments_styles_dimensions_merges_and_sheets(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    initial, updated = ENTITY_CASES["resource"]
    _write_table(path, "resource", [("gold", initial)], extra_header="notes")
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["E4"] = "=1+1"
    sheet["E4"].comment = Comment("keep me", "tester")
    sheet["B4"].font = Font(bold=True, color="ABCDEF")
    sheet.column_dimensions["E"].width = 27.5
    sheet.row_dimensions[4].height = 31
    sheet.merge_cells("G7:H7")
    sheet["G7"] = "merged"
    other = workbook.create_sheet("Keep")
    other["A1"] = "=SUM(1,2)"
    other.sheet_properties.tabColor = "00FF00"
    workbook.save(path)
    workbook.close()

    assert upsert_workbook_entity(path, _change("resource", "gold", updated))

    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    assert sheet["E4"].value == "=1+1"
    assert sheet["E4"].comment.text == "keep me"
    assert sheet["B4"].font.bold is True
    assert sheet.column_dimensions["E"].width == 27.5
    assert sheet.row_dimensions[4].height == 31
    assert "G7:H7" in {str(item) for item in sheet.merged_cells.ranges}
    assert workbook["Keep"]["A1"].value == "=SUM(1,2)"
    assert workbook["Keep"].sheet_properties.tabColor.rgb.endswith("00FF00")
    workbook.close()


def test_update_preserves_unrelated_rich_text_runs_and_inline_fonts(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    initial, updated = ENTITY_CASES["resource"]
    _write_table(path, "resource", [("gold", initial)], extra_header="notes")
    workbook = load_workbook(path)
    workbook.active["E4"] = CellRichText(
        "plain ",
        TextBlock(InlineFont(b=True, color="FF112233"), "bold"),
        TextBlock(InlineFont(i=True, color="FF445566"), " italic"),
    )
    workbook.save(path)
    workbook.close()

    assert upsert_workbook_entity(path, _change("resource", "gold", updated))

    workbook = load_workbook(path, data_only=False, rich_text=True)
    value = workbook.active["E4"].value
    assert isinstance(value, CellRichText)
    assert str(value) == "plain bold italic"
    assert isinstance(value[1], TextBlock)
    assert value[1].text == "bold"
    assert value[1].font.b is True
    assert value[1].font.color.rgb == "FF112233"
    assert isinstance(value[2], TextBlock)
    assert value[2].text == " italic"
    assert value[2].font.i is True
    assert value[2].font.color.rgb == "FF445566"
    workbook.close()


@pytest.mark.parametrize(
    ("mutation", "reason"),
    [
        ("duplicate_marker", "duplicate_marker"),
        ("split_marker_columns", "incoherent_markers"),
        ("duplicate_header", "duplicate_header"),
        ("missing_header", "missing_column"),
        ("wrong_type", "wrong_column_type"),
    ],
)
def test_malformed_marker_and_header_schemas_are_structured(
    tmp_path: Path, mutation: str, reason: str
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    workbook = load_workbook(path)
    sheet = workbook.active
    if mutation == "duplicate_marker":
        sheet["A8"] = "##var"
    elif mutation == "split_marker_columns":
        sheet["C2"] = sheet["A2"].value
        sheet["A2"] = None
    elif mutation == "duplicate_header":
        sheet["E1"] = "name"
    elif mutation == "missing_header":
        sheet["D1"] = "unknown"
    elif mutation == "wrong_type":
        sheet["D3"] = "int"
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == reason
    json.dumps(dict(caught.value.details))


def test_missing_corrupt_wrong_named_and_oversized_sources_are_structured(
    tmp_path: Path,
) -> None:
    missing = tmp_path / "resources.xlsx"
    with pytest.raises(AuthoringError) as missing_error:
        inspect_table(missing)
    assert missing_error.value.details["reason"] == "read_error"

    corrupt = tmp_path / "resources.xlsx"
    corrupt.write_bytes(b"not a workbook")
    with pytest.raises(AuthoringError) as corrupt_error:
        inspect_table(corrupt)
    assert corrupt_error.value.details["reason"] == "corrupt_workbook"

    with ZipFile(corrupt, "w") as archive:
        archive.writestr("[Content_Types].xml", "<broken")
    with pytest.raises(AuthoringError) as broken_xml_error:
        inspect_table(corrupt)
    assert broken_xml_error.value.code == "invalid_workbook_source"
    assert broken_xml_error.value.details["reason"] == "corrupt_workbook"

    wrong = tmp_path / "unknown.xlsx"
    wrong.write_bytes(corrupt.read_bytes())
    with pytest.raises(AuthoringError) as wrong_error:
        inspect_table(wrong)
    assert wrong_error.value.details["reason"] == "unknown_table"

    huge = tmp_path / "resources.xlsx"
    _write_table(huge, "resource")
    workbook = load_workbook(huge)
    workbook.active.cell(100_001, 2, "too far")
    workbook.save(huge)
    workbook.close()
    with pytest.raises(AuthoringError) as huge_error:
        inspect_table(huge)
    assert huge_error.value.details["reason"] == "dimension_limit"
    assert huge_error.value.details["limit_rows"] == 100_000


def test_worksheet_budget_is_enforced_before_openpyxl_materializes_cells(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    workbook = load_workbook(path)
    sheet = workbook.active
    for column in range(1, 33):
        sheet.cell(10, column, f"cell-{column}")
    workbook.save(path)
    workbook.close()

    monkeypatch.setattr("igess.authoring.workbook_source._MAX_CELLS", 20)

    def must_not_materialize(*args: Any, **kwargs: Any) -> NoReturn:
        del args, kwargs
        raise AssertionError("openpyxl was called before worksheet preflight")

    monkeypatch.setattr(
        "igess.authoring.workbook_source.load_workbook",
        must_not_materialize,
    )
    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "dimension_limit"
    assert caught.value.details["limit_cells"] == 20


def test_streaming_preflight_counts_cells_even_when_xml_dimension_lies(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    workbook = load_workbook(path)
    sheet = workbook.active
    for column in range(1, 33):
        sheet.cell(10, column, f"cell-{column}")
    workbook.save(path)
    workbook.close()

    rewritten = path.with_name("rewritten.xlsx")
    replaced_dimension = False
    with ZipFile(path) as source, ZipFile(rewritten, "w") as target:
        for member in source.infolist():
            content = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                changed = re.sub(
                    rb'<dimension ref="[^"]+"\s*/>',
                    b'<dimension ref="A1"/>',
                    content,
                    count=1,
                )
                changed = changed.replace(b'<row r="10">', b'<row r="1">')
                changed = re.sub(
                    rb'<c r="[A-Z]+10"',
                    b'<c r="A1"',
                    changed,
                )
                replaced_dimension = changed != content
                content = changed
            target.writestr(member, content)
    assert replaced_dimension is True
    os.replace(rewritten, path)
    monkeypatch.setattr("igess.authoring.workbook_source._MAX_CELLS", 20)

    def must_not_materialize(*args: Any, **kwargs: Any) -> NoReturn:
        del args, kwargs
        raise AssertionError("openpyxl was called before worksheet cell counting")

    monkeypatch.setattr(
        "igess.authoring.workbook_source.load_workbook",
        must_not_materialize,
    )
    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.details["reason"] == "dimension_limit"
    assert caught.value.details["actual_cells"] == 21


def test_inspection_loads_the_preflighted_snapshot_if_path_changes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _path_for(tmp_path, "resource")
    replacement = tmp_path / "replacement.xlsx"
    initial = ENTITY_CASES["resource"][0]
    _write_table(path, "resource", [("gold", initial)])
    _write_table(replacement, "resource", [("fish", initial)])
    replacement_bytes = replacement.read_bytes()
    real_load = load_workbook
    swapped = False

    def swap_path_before_load(source: Any, *args: Any, **kwargs: Any) -> Any:
        nonlocal swapped
        if not swapped:
            swapped = True
            path.write_bytes(replacement_bytes)
        return real_load(source, *args, **kwargs)

    monkeypatch.setattr(
        "igess.authoring.workbook_source.load_workbook",
        swap_path_before_load,
    )

    inspected = inspect_table(path)

    assert swapped is True
    assert [record.entity_id for record in inspected.records] == ["gold"]


def test_upsert_does_not_overwrite_path_changed_after_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = _path_for(tmp_path, "resource")
    replacement = tmp_path / "replacement.xlsx"
    _write_table(path, "resource")
    _write_table(
        replacement,
        "resource",
        [("external", ENTITY_CASES["resource"][0])],
    )
    replacement_bytes = replacement.read_bytes()
    real_load = load_workbook
    swapped = False

    def swap_path_before_load(source: Any, *args: Any, **kwargs: Any) -> Any:
        nonlocal swapped
        if not swapped:
            swapped = True
            path.write_bytes(replacement_bytes)
        return real_load(source, *args, **kwargs)

    monkeypatch.setattr(
        "igess.authoring.workbook_source.load_workbook",
        swap_path_before_load,
    )

    with pytest.raises(AuthoringError) as caught:
        upsert_workbook_entity(
            path,
            _change("resource", "gold", ENTITY_CASES["resource"][0]),
        )

    assert caught.value.code == "workbook_write_failed"
    assert caught.value.details["reason"] == "source_changed"
    assert path.read_bytes() == replacement_bytes
    assert _temp_files(path) == []


@pytest.mark.parametrize(
    ("part", "feature"),
    [
        ("xl/externalLinks/externalLink1.xml", "external_links"),
        ("xl/vbaProject.bin", "vba_macros"),
    ],
)
def test_lossy_package_parts_are_rejected_before_inspect_or_mutation(
    tmp_path: Path,
    part: str,
    feature: str,
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    with ZipFile(path, "a") as archive:
        archive.writestr(part, b"synthetic unsupported package part")
    before = path.read_bytes()

    with pytest.raises(AuthoringError) as inspect_error:
        inspect_table(path)
    assert inspect_error.value.code == "invalid_workbook_source"
    assert inspect_error.value.details["reason"] == "unsupported_package_part"
    assert inspect_error.value.details["feature"] == feature
    assert inspect_error.value.details["part"] == part

    with pytest.raises(AuthoringError) as upsert_error:
        upsert_workbook_entity(
            path,
            _change("resource", "gold", ENTITY_CASES["resource"][0]),
        )
    assert upsert_error.value.details["reason"] == "unsupported_package_part"
    assert path.read_bytes() == before


@pytest.mark.parametrize(
    ("entity", "cell", "raw_value", "field"),
    [
        ("resource", "B4", 123, "id"),
        ("resource", "C4", 123, "name"),
        ("resource", "D4", True, "dimension"),
        ("resource", "C4", date(2026, 7, 15), "name"),
        ("resource", "C4", "=1+1", "name"),
        ("constant", "C4", 1.25, "value"),
    ],
)
def test_source_cells_keep_exact_types_instead_of_string_coercion(
    tmp_path: Path,
    entity: str,
    cell: str,
    raw_value: Any,
    field: str,
) -> None:
    path = _path_for(tmp_path, entity)
    _write_table(path, entity, [("entry", ENTITY_CASES[entity][0])])
    workbook = load_workbook(path)
    workbook.active[cell] = raw_value
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "invalid_entity_row"
    assert caught.value.details["field"] == field
    assert caught.value.details["row"] == 4
    CommandResponse(
        command="inspect",
        ok=False,
        code=caught.value.code,
        message=caught.value.message,
        details=caught.value.details,
    ).to_json()


@pytest.mark.parametrize(
    ("entity", "field"),
    [
        ("resource", "id"),
        ("resource", "name"),
        ("activity", "unlock_condition"),
        ("prestige_layer", "reset_resources"),
    ],
)
def test_excel_error_cells_are_never_accepted_as_literal_strings(
    tmp_path: Path,
    entity: str,
    field: str,
) -> None:
    path = _path_for(tmp_path, entity)
    _write_table(path, entity, [("entry", ENTITY_CASES[entity][0])])
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = {
        sheet.cell(1, column).value: column
        for column in range(2, sheet.max_column + 1)
    }
    cell = sheet.cell(4, headers[field], "#N/A")
    assert cell.data_type == "e"
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "invalid_entity_row"
    assert caught.value.details["row"] == 4
    assert caught.value.details["field"] == field
    assert caught.value.details["source_type"] == "e"


@pytest.mark.parametrize("raw_value", ["gold;;ore", ";gold", "gold;", 123, True])
def test_prestige_list_encoding_rejects_empty_tokens_and_non_strings(
    tmp_path: Path,
    raw_value: Any,
) -> None:
    path = _path_for(tmp_path, "prestige_layer")
    _write_table(
        path,
        "prestige_layer",
        [("entry", ENTITY_CASES["prestige_layer"][0])],
    )
    workbook = load_workbook(path)
    headers = {
        workbook.active.cell(1, column).value: column
        for column in range(2, workbook.active.max_column + 1)
    }
    workbook.active.cell(4, headers["reset_resources"], raw_value)
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "invalid_entity_row"
    assert caught.value.details["field"] == "reset_resources"


def test_integer_decimal_cells_remain_exact_and_supported(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "constant")
    _write_table(path, "constant", [("starting_gold", {"value": 100})])

    inspected = inspect_table(path)

    assert _fields(inspected.records[0]) == {"value": "100"}


def test_orphan_schema_values_are_rejected(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    workbook = load_workbook(path)
    workbook.active["C4"] = "No id"
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.details["reason"] == "missing_id"
    assert caught.value.details["row"] == 4


def test_invalid_existing_entity_is_reported_as_a_row_aware_source_error(
    tmp_path: Path,
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource", [("bad", {"name": "Bad", "dimension": "valid"})])
    workbook = load_workbook(path)
    workbook.active["D4"] = "not valid id!"
    workbook.save(path)
    workbook.close()

    with pytest.raises(AuthoringError) as caught:
        inspect_table(path)

    assert caught.value.code == "invalid_workbook_source"
    assert caught.value.details["reason"] == "invalid_entity_row"
    assert caught.value.details["entity"] == "resource"
    assert caught.value.details["id"] == "bad"
    assert caught.value.details["row"] == 4
    assert caught.value.details["field"] == "dimension"


def test_write_validates_change_entity_matches_path_and_current_rows(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    with pytest.raises(AuthoringError) as mismatch:
        upsert_workbook_entity(
            path, _change("constant", "starting_gold", {"value": "1"})
        )
    assert mismatch.value.details["reason"] == "entity_path_mismatch"

    _write_table(path, "resource", [("bad", {"name": "Bad", "dimension": "space"})])
    workbook = load_workbook(path)
    workbook.active["D4"] = "not valid id!"
    workbook.save(path)
    workbook.close()
    before = path.read_bytes()
    with pytest.raises(AuthoringError):
        upsert_workbook_entity(
            path, _change("resource", "gold", ENTITY_CASES["resource"][0])
        )
    assert path.read_bytes() == before


def _temp_files(path: Path) -> list[Path]:
    return list(path.parent.glob(f".{path.name}.*.tmp.xlsx"))


def test_reopen_failure_leaves_original_and_cleans_temp(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    before = path.read_bytes()
    real_load = load_workbook
    load_calls = 0

    def fail_temp(source: Any, *args: Any, **kwargs: Any) -> Any:
        nonlocal load_calls
        load_calls += 1
        if load_calls == 2:
            raise OSError("reopen failed")
        return real_load(source, *args, **kwargs)

    monkeypatch.setattr("igess.authoring.workbook_source.load_workbook", fail_temp)
    with pytest.raises(AuthoringError) as caught:
        upsert_workbook_entity(
            path, _change("resource", "gold", ENTITY_CASES["resource"][0])
        )

    assert caught.value.code == "workbook_write_failed"
    assert caught.value.details["reason"] == "reload_error"
    assert path.read_bytes() == before
    assert _temp_files(path) == []


def test_replace_failure_leaves_original_and_cleans_temp(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    before = path.read_bytes()

    def fail_replace(source: Any, target: Any) -> None:
        del source, target
        raise OSError("replace failed")

    monkeypatch.setattr("igess.authoring.workbook_source.os.replace", fail_replace)
    with pytest.raises(AuthoringError) as caught:
        upsert_workbook_entity(
            path, _change("resource", "gold", ENTITY_CASES["resource"][0])
        )

    assert caught.value.code == "workbook_write_failed"
    assert caught.value.details["reason"] == "replace_error"
    assert path.read_bytes() == before
    assert _temp_files(path) == []


class _ReplaceSignal(BaseException):
    pass


@pytest.mark.parametrize("error_type", [OSError, _ReplaceSignal, KeyboardInterrupt])
def test_replace_that_installs_exact_bytes_then_raises_is_success(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    error_type: type[BaseException],
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    real_replace = os.replace

    def replace_then_raise(source: Any, target: Any) -> NoReturn:
        real_replace(source, target)
        raise error_type("after success")

    monkeypatch.setattr(
        "igess.authoring.workbook_source.os.replace", replace_then_raise
    )
    assert upsert_workbook_entity(
        path, _change("resource", "gold", ENTITY_CASES["resource"][0])
    ) is True
    assert _fields(inspect_table(path).records[0]) == ENTITY_CASES["resource"][0]
    assert _temp_files(path) == []


def test_base_exception_before_replace_propagates_without_mutation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    before = path.read_bytes()

    def interrupt_save(self: Any, filename: Any) -> NoReturn:
        del self, filename
        raise KeyboardInterrupt("before replace")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", interrupt_save)
    with pytest.raises(KeyboardInterrupt):
        upsert_workbook_entity(
            path, _change("resource", "gold", ENTITY_CASES["resource"][0])
        )
    assert path.read_bytes() == before
    assert _temp_files(path) == []


def test_atomic_write_applies_only_source_permission_bits_to_temp(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    expected_mode = stat.S_IMODE(path.stat().st_mode)
    calls: list[tuple[Path, int]] = []
    real_chmod = os.chmod

    def record_chmod(target: str | os.PathLike[str], mode: int) -> None:
        calls.append((Path(target), mode))
        real_chmod(target, mode)

    monkeypatch.setattr("igess.authoring.workbook_source.os.chmod", record_chmod)
    assert upsert_workbook_entity(
        path, _change("resource", "gold", ENTITY_CASES["resource"][0])
    )
    assert len(calls) == 1
    assert calls[0][0].parent == path.parent
    assert calls[0][0] != path
    assert calls[0][1] == expected_mode


@pytest.mark.skipif(os.name != "posix", reason="POSIX permission contract")
def test_atomic_write_preserves_posix_permission_mode(tmp_path: Path) -> None:
    path = _path_for(tmp_path, "resource")
    _write_table(path, "resource")
    path.chmod(0o640)
    assert upsert_workbook_entity(
        path, _change("resource", "gold", ENTITY_CASES["resource"][0])
    )
    assert stat.S_IMODE(path.stat().st_mode) == 0o640
