from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class TableRegistration:
    table: str
    path: str
    mode: str
    key: str


def export_registered_workbooks(datas_dir: str | Path, output_dir: str | Path) -> list[Path]:
    datas_path = Path(datas_dir)
    output_path = Path(output_dir)
    registrations = _read_registrations(datas_path / "__tables__.xlsx")
    expected_names = {f"{registration.table}.json" for registration in registrations}
    output_path.mkdir(parents=True, exist_ok=True)
    for stale_path in output_path.glob("*.json"):
        if stale_path.name not in expected_names:
            stale_path.unlink()
    written: list[Path] = []
    for registration in registrations:
        if registration.mode != "map":
            raise ValueError(
                f"table '{registration.table}' unsupported mode '{registration.mode}'"
            )
        if registration.key != "id":
            raise ValueError(
                f"table '{registration.table}' unsupported key '{registration.key}'"
            )
        workbook_path = datas_path / registration.path
        target_path = output_path / f"{registration.table}.json"
        export_workbook_to_json(workbook_path, target_path, table_name=registration.table)
        written.append(target_path)
    return written


def export_workbook_to_json(
    workbook_path: str | Path, output_path: str | Path, table_name: str | None = None
) -> None:
    workbook_path = Path(workbook_path)
    output_path = Path(output_path)
    table = table_name or workbook_path.stem
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    _validate_marker_rows(workbook_path, sheet)
    headers = [cell.value for cell in sheet[1]][1:]
    types = [cell.value for cell in sheet[3]][1:]
    rows: list[dict[str, Any]] = []
    for row_index, values in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        data_values = values[1:]
        if all(value in (None, "") for value in data_values):
            continue
        payload = {}
        for header, field_type, value in zip(headers, types, data_values):
            if header in (None, ""):
                continue
            payload[str(header)] = _export_value(value, str(field_type or "string"))
        payload["_source"] = {
            "table": table,
            "workbook": workbook_path.name,
            "row": row_index,
        }
        rows.append(payload)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def _read_registrations(path: Path) -> list[TableRegistration]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    _validate_marker_rows(path, sheet)
    headers = [cell.value for cell in sheet[1]][1:]
    registrations: list[TableRegistration] = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        data = dict(zip(headers, values[1:]))
        if not data.get("table"):
            continue
        registrations.append(
            TableRegistration(
                table=str(data["table"]),
                path=str(data["path"]),
                mode=str(data.get("mode", "map")),
                key=str(data.get("key", "id")),
            )
        )
    return registrations


def _validate_marker_rows(path: Path, sheet) -> None:
    if sheet["A1"].value != "##var":
        raise ValueError(f"{path} is missing ##var marker in A1")
    if sheet["A2"].value != "##":
        raise ValueError(f"{path} is missing ## marker in A2")
    if sheet["A3"].value != "##type":
        raise ValueError(f"{path} is missing ##type marker in A3")


def _export_value(value: Any, field_type: str) -> Any:
    if value is None:
        value = ""
    if field_type.startswith("(list"):
        return [part for part in str(value).split(";") if part]
    return str(value)
