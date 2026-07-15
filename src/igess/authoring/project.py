"""Authoring project discovery, canonical paths, and source digests."""

from __future__ import annotations

from contextlib import ExitStack, contextmanager
from dataclasses import dataclass, field
import hashlib
import os
from pathlib import Path
import stat
from typing import Any, BinaryIO, Iterator, Literal, NoReturn

from openpyxl import load_workbook

from .response import AuthoringError


_PROJECT_PATHS = (
    ("economy.yaml", "file", "project config", "project_config"),
    ("Datas", "directory", "source tables", "source_tables"),
    ("luban_exports", "directory", "runtime exports", "runtime_exports"),
)

_DIGEST_CHUNK_SIZE = 1024 * 1024


@dataclass(frozen=True, slots=True)
class _SourceSnapshot:
    kind: Literal["config", "registry", "workbook"]
    path: Path
    identity: os.stat_result | None
    registry: Path | None = None
    row: int | None = None
    registration_path: str | None = None


@dataclass(frozen=True, slots=True)
class _OpenedSource:
    snapshot: _SourceSnapshot
    path: Path
    handle: BinaryIO
    identity: os.stat_result


@dataclass(frozen=True, slots=True)
class AuthoringProject:
    """The canonical locations belonging to one incremental-authoring project."""

    root: Path
    config: Path = field(init=False)
    datas: Path = field(init=False)
    exports: Path = field(init=False)
    runs: Path = field(init=False)
    legacy_runs: Path = field(init=False)
    reports: Path = field(init=False)
    changes: Path = field(init=False)
    transactions: Path = field(init=False)
    lock: Path = field(init=False)

    def __post_init__(self) -> None:
        if not isinstance(self.root, (str, os.PathLike)):
            raise TypeError("AuthoringProject root must be a path-like value")
        try:
            canonical_root = Path(self.root).expanduser().resolve()
        except (OSError, RuntimeError, ValueError) as error:
            raise AuthoringError(
                "project_root_invalid",
                "Authoring project root could not be resolved",
                {
                    "error_type": type(error).__name__,
                    "path": str(self.root),
                    "reason": "resolve_error",
                    "role": "project root",
                },
            ) from None

        discovered = {
            name: _require_project_path(canonical_root, name, expected, role, code_prefix)
            for name, expected, role, code_prefix in _PROJECT_PATHS
        }
        metadata = canonical_root / ".igess"
        object.__setattr__(self, "root", canonical_root)
        object.__setattr__(self, "config", discovered["economy.yaml"])
        object.__setattr__(self, "datas", discovered["Datas"])
        object.__setattr__(self, "exports", discovered["luban_exports"])
        object.__setattr__(self, "runs", canonical_root / "runs")
        object.__setattr__(self, "legacy_runs", metadata / "runs")
        object.__setattr__(self, "reports", canonical_root / "reports")
        object.__setattr__(self, "changes", canonical_root / "changes")
        object.__setattr__(self, "transactions", metadata / "transactions")
        object.__setattr__(self, "lock", metadata / "model.lock")

    @classmethod
    def discover(cls, root: str | Path) -> AuthoringProject:
        """Discover a project without recursively searching below ``root``."""

        return cls(Path(root))

    def read_run_roots(self) -> list[Path]:
        """Return readable run registries, preferring the modern root."""

        result: list[Path] = []
        identities: set[str] = set()
        for candidate in (self.runs, self.legacy_runs):
            try:
                if not candidate.is_dir():
                    continue
                resolved = candidate.resolve(strict=True)
            except (OSError, RuntimeError, ValueError):
                continue
            identity = os.path.normcase(str(resolved))
            if identity in identities or any(_same_file(candidate, existing) for existing in result):
                continue
            identities.add(identity)
            result.append(candidate)
        return result

    def model_digest(self) -> str:
        """Hash exactly the config, registry, and registered source workbooks."""

        config = _current_required_path(
            self,
            "economy.yaml",
            "file",
            "project config",
            "project_config",
        )
        datas = _current_required_path(
            self,
            "Datas",
            "directory",
            "source tables",
            "source_tables",
        )
        registry = _resolve_registry(datas / "__tables__.xlsx", datas)
        config_snapshot = _snapshot_source("config", config)
        registry_snapshot = _snapshot_source("registry", registry, registry=registry)

        with ExitStack() as stack:
            opened_config = stack.enter_context(
                _open_validated_source(
                    config_snapshot,
                    root=self.root,
                    boundary=self.root,
                    direct=True,
                )
            )
            opened_registry = stack.enter_context(
                _open_validated_source(
                    registry_snapshot,
                    root=self.root,
                    boundary=datas,
                    direct=True,
                )
            )
            registrations = _read_registration_paths(opened_registry.handle, registry)
            source_snapshots = _validate_registration_paths(datas, registry, registrations)
            opened_sources = [
                stack.enter_context(
                    _open_validated_source(
                        snapshot,
                        root=self.root,
                        boundary=datas,
                        direct=False,
                    )
                )
                for snapshot in source_snapshots
            ]
            _reject_duplicate_opened_sources(opened_sources, registry)

            digest = hashlib.sha256()
            all_sources = [opened_config, opened_registry, *opened_sources]
            for source in sorted(
                all_sources,
                key=lambda item: _root_relative_posix(self.root, item.path),
            ):
                relative = _root_relative_posix(self.root, source.path)
                digest.update(relative.encode("utf-8"))
                digest.update(b"\0")
                _stream_opened_source(digest, source)
            return f"sha256:{digest.hexdigest()}"


def _require_project_path(
    root: Path,
    name: str,
    expected: str,
    role: str,
    code_prefix: str,
) -> Path:
    path = root / name
    try:
        resolved = path.resolve(strict=True)
    except FileNotFoundError:
        _project_path_error(
            f"{code_prefix}_missing",
            f"Required {role} is missing: {path}",
            path,
            expected,
            role,
            "missing",
        )
    except (OSError, RuntimeError, ValueError) as error:
        _project_path_error(
            f"{code_prefix}_inaccessible",
            f"Required {role} could not be resolved: {path}",
            path,
            expected,
            role,
            "resolve_error",
            error_type=type(error).__name__,
        )

    if not resolved.is_relative_to(root):
        _project_path_error(
            f"{code_prefix}_unsafe",
            f"Required {role} resolves outside the project root: {path}",
            path,
            expected,
            role,
            "outside_root",
            resolved_path=str(resolved),
        )
    if resolved.parent != root:
        _project_path_error(
            f"{code_prefix}_unsafe",
            f"Required {role} must resolve to a direct child: {path}",
            path,
            expected,
            role,
            "not_direct_child",
            resolved_path=str(resolved),
        )

    try:
        mode = resolved.stat().st_mode
    except OSError as error:
        _project_path_error(
            f"{code_prefix}_inaccessible",
            f"Required {role} could not be inspected: {path}",
            path,
            expected,
            role,
            "access_error",
            error_type=type(error).__name__,
        )
    matches = stat.S_ISREG(mode) if expected == "file" else stat.S_ISDIR(mode)
    if not matches:
        _project_path_error(
            f"{code_prefix}_wrong_type",
            f"Required {role} is not a {expected}: {path}",
            path,
            expected,
            role,
            "wrong_type",
            resolved_path=str(resolved),
        )
    return resolved


def _current_required_path(
    project: AuthoringProject,
    name: str,
    expected: str,
    role: str,
    code_prefix: str,
) -> Path:
    current = _require_project_path(project.root, name, expected, role, code_prefix)
    original = getattr(project, {"economy.yaml": "config", "Datas": "datas"}[name])
    if current != original:
        _project_path_error(
            f"{code_prefix}_unsafe",
            f"Required {role} was retargeted after project discovery",
            project.root / name,
            expected,
            role,
            "path_retargeted",
            resolved_path=str(current),
        )
    return current


def _project_path_error(
    code: str,
    message: str,
    path: Path,
    expected: str,
    role: str,
    reason: str,
    **details: str,
) -> NoReturn:
    raise AuthoringError(
        code,
        message,
        {
            "expected": expected,
            "path": str(path),
            "reason": reason,
            "role": role,
            **details,
        },
    )


def _same_file(first: Path, second: Path) -> bool:
    try:
        return os.path.samefile(first, second)
    except OSError:
        return False


def _snapshot_source(
    kind: Literal["config", "registry", "workbook"],
    path: Path,
    *,
    registry: Path | None = None,
    row: int | None = None,
    registration_path: str | None = None,
) -> _SourceSnapshot:
    partial = _SourceSnapshot(
        kind=kind,
        path=path,
        identity=None,
        registry=registry,
        row=row,
        registration_path=registration_path,
    )
    try:
        identity = path.stat()
    except (OSError, MemoryError, ValueError) as error:
        _source_error(
            partial,
            "A model source could not be inspected",
            "source_access_error",
            error_type=type(error).__name__,
        )
    if not stat.S_ISREG(identity.st_mode):
        _source_error(
            partial,
            "A model source is not a regular file",
            "source_wrong_type",
        )
    return _SourceSnapshot(
        kind=kind,
        path=path,
        identity=identity,
        registry=registry,
        row=row,
        registration_path=registration_path,
    )


def _open_binary(path: Path) -> BinaryIO:
    """Open a source for identity validation before any content is consumed."""

    return path.open("rb")


@contextmanager
def _open_validated_source(
    snapshot: _SourceSnapshot,
    *,
    root: Path,
    boundary: Path,
    direct: bool,
) -> Iterator[_OpenedSource]:
    handle: BinaryIO | None = None
    try:
        try:
            handle = _open_binary(snapshot.path)
            opened_identity = os.fstat(handle.fileno())
        except (OSError, MemoryError, TypeError, ValueError) as error:
            _source_error(
                snapshot,
                "A model source could not be opened",
                "source_open_error",
                error_type=type(error).__name__,
            )

        if snapshot.identity is None or not os.path.samestat(snapshot.identity, opened_identity):
            _source_error(
                snapshot,
                "A model source changed between validation and opening",
                "path_identity_changed",
            )
        if not stat.S_ISREG(opened_identity.st_mode):
            _source_error(
                snapshot,
                "An opened model source is not a regular file",
                "source_wrong_type",
            )

        try:
            resolved_root = root.resolve(strict=True)
            resolved_boundary = boundary.resolve(strict=True)
            resolved_path = snapshot.path.resolve(strict=True)
            path_identity = resolved_path.stat()
        except (OSError, RuntimeError, ValueError) as error:
            _source_error(
                snapshot,
                "A model source boundary could not be verified after opening",
                "post_open_validation_error",
                error_type=type(error).__name__,
            )
        if resolved_root != root or resolved_boundary != boundary:
            _source_error(
                snapshot,
                "A model source boundary changed while opening",
                "boundary_identity_changed",
                resolved_path=str(resolved_path),
            )
        if boundary != root and boundary.parent != root:
            _source_error(
                snapshot,
                "A model source boundary is no longer a direct project child",
                "boundary_not_direct_child",
                resolved_path=str(resolved_boundary),
            )
        if not resolved_path.is_relative_to(resolved_boundary):
            _source_error(
                snapshot,
                "An opened model source resolves outside its allowed boundary",
                "outside_root",
                resolved_path=str(resolved_path),
            )
        if direct and resolved_path.parent != resolved_boundary:
            _source_error(
                snapshot,
                "An opened model source is no longer a direct child",
                "not_direct_child",
                resolved_path=str(resolved_path),
            )
        if not os.path.samestat(opened_identity, path_identity):
            _source_error(
                snapshot,
                "A model source path changed after opening",
                "path_identity_changed",
                resolved_path=str(resolved_path),
            )

        yield _OpenedSource(snapshot, resolved_path, handle, opened_identity)
    finally:
        if handle is not None:
            try:
                handle.close()
            except (OSError, MemoryError, ValueError) as error:
                _source_error(
                    snapshot,
                    "A model source handle could not be closed safely",
                    "source_close_error",
                    error_type=type(error).__name__,
                )


def _source_error(
    snapshot: _SourceSnapshot,
    message: str,
    reason: str,
    **details: str,
) -> NoReturn:
    if snapshot.kind == "config":
        unsafe_reasons = {
            "boundary_identity_changed",
            "boundary_not_direct_child",
            "not_direct_child",
            "outside_root",
            "path_identity_changed",
            "post_open_validation_error",
        }
        _project_path_error(
            "project_config_unsafe" if reason in unsafe_reasons else "project_config_unreadable",
            message,
            snapshot.path,
            "file",
            "project config",
            reason,
            **details,
        )

    registry = snapshot.registry or snapshot.path
    source_details: dict[str, str | int] = dict(details)
    if snapshot.kind == "workbook":
        source_details["source_path"] = str(snapshot.path)
    if snapshot.row is not None:
        source_details["row"] = snapshot.row
    if snapshot.registration_path is not None:
        source_details["registration_path"] = snapshot.registration_path
    _registry_error(message, reason, registry, **source_details)


def _reject_duplicate_opened_sources(sources: list[_OpenedSource], registry: Path) -> None:
    for index, source in enumerate(sources):
        for earlier in sources[:index]:
            if os.path.samestat(source.identity, earlier.identity):
                _registry_error(
                    "A workbook file is registered more than once",
                    "duplicate_registration_path",
                    registry,
                    source_path=str(source.path),
                )


def _resolve_registry(path: Path, datas: Path) -> Path:
    try:
        registry = path.resolve(strict=True)
    except FileNotFoundError:
        _registry_error("Source registry is missing", "registry_missing", path)
    except (OSError, RuntimeError, ValueError) as error:
        _registry_error(
            "Source registry could not be resolved",
            "registry_inaccessible",
            path,
            error_type=type(error).__name__,
        )
    if not registry.is_relative_to(datas):
        _registry_error(
            "Source registry resolves outside Datas",
            "unsafe_registry_path",
            path,
            resolved_path=str(registry),
        )
    if registry.parent != datas:
        _registry_error(
            "Source registry must resolve directly below Datas",
            "unsafe_registry_path",
            path,
            resolved_path=str(registry),
        )
    try:
        if not stat.S_ISREG(registry.stat().st_mode):
            _registry_error("Source registry is not a file", "registry_wrong_type", path)
    except OSError as error:
        _registry_error(
            "Source registry could not be inspected",
            "registry_inaccessible",
            path,
            error_type=type(error).__name__,
        )
    return registry


def _read_registration_paths(
    registry_handle: BinaryIO,
    registry: Path,
) -> list[tuple[int, Any]]:
    workbook = None
    try:
        registry_handle.seek(0)
        workbook = load_workbook(registry_handle, read_only=True, data_only=True)
        sheet = workbook.active
        if (
            sheet["A1"].value != "##var"
            or sheet["A2"].value != "##"
            or sheet["A3"].value != "##type"
        ):
            _registry_error(
                "Source registry has invalid Luban marker rows",
                "malformed_registry",
                registry,
            )
        headers = [cell.value for cell in sheet[1]][1:]
        nonempty_headers = [header for header in headers if header not in (None, "")]
        duplicate_headers = _duplicate_values(nonempty_headers)
        if duplicate_headers:
            _registry_error(
                "Source registry contains duplicate headers",
                "malformed_registry",
                registry,
                header_issue="duplicate_header",
                header=str(duplicate_headers[0]),
            )
        missing_headers = [header for header in ("table", "path") if headers.count(header) != 1]
        if missing_headers:
            _registry_error(
                "Source registry is missing a required header",
                "malformed_registry",
                registry,
                header_issue="missing_required_header",
                header=missing_headers[0],
            )
        invalid_headers = [header for header in nonempty_headers if not isinstance(header, str)]
        if invalid_headers:
            _registry_error(
                "Source registry contains a non-string header",
                "malformed_registry",
                registry,
                header_issue="invalid_header",
                header_type=type(invalid_headers[0]).__name__,
            )

        path_index = headers.index("path") + 1
        table_index = headers.index("table") + 1

        registrations: list[tuple[int, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=4, values_only=True),
            start=4,
        ):
            if values[table_index] in (None, ""):
                continue
            registrations.append((row_number, values[path_index]))
        return registrations
    except AuthoringError:
        raise
    except Exception as error:
        _registry_error(
            "Source registry could not be read",
            "malformed_registry",
            registry,
            error_type=type(error).__name__,
        )
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception as error:
                _registry_error(
                    "Source registry could not be closed safely",
                    "registry_close_error",
                    registry,
                    error_type=type(error).__name__,
                )


def _duplicate_values(values: list[Any]) -> list[Any]:
    duplicates: list[Any] = []
    seen: list[Any] = []
    for value in values:
        if value in seen:
            if value not in duplicates:
                duplicates.append(value)
        else:
            seen.append(value)
    return duplicates


def _validate_registration_paths(
    datas_root: Path,
    registry: Path,
    registrations: list[tuple[int, Any]],
) -> list[_SourceSnapshot]:
    identities: set[str] = set()
    result: list[_SourceSnapshot] = []
    for row, value in registrations:
        if not isinstance(value, str) or not value.strip():
            _registry_error(
                "A source registration has no workbook path",
                "missing_registration_path",
                registry,
                row=row,
            )
        try:
            relative = Path(value)
        except (TypeError, ValueError) as error:
            _registry_error(
                "A source registration path is invalid",
                "unsafe_registration_path",
                registry,
                row=row,
                registration_path=str(value),
                error_type=type(error).__name__,
            )
        if relative.is_absolute() or relative.anchor or ".." in relative.parts:
            _registry_error(
                "A source registration path must stay below Datas",
                "unsafe_registration_path",
                registry,
                row=row,
                registration_path=value,
            )

        try:
            source = (datas_root / relative).resolve(strict=True)
            source.relative_to(datas_root)
        except FileNotFoundError:
            _registry_error(
                "A registered source workbook is missing",
                "registered_source_missing",
                registry,
                row=row,
                registration_path=value,
                source_path=str(datas_root / relative),
            )
        except (OSError, RuntimeError, ValueError) as error:
            _registry_error(
                "A source registration path escapes Datas",
                "unsafe_registration_path",
                registry,
                row=row,
                registration_path=value,
                error_type=type(error).__name__,
            )

        identity = os.path.normcase(str(source))
        if identity in identities:
            _registry_error(
                "A workbook path is registered more than once",
                "duplicate_registration_path",
                registry,
                row=row,
                registration_path=value,
            )
        identities.add(identity)
        try:
            source_identity = source.stat()
        except OSError as error:
            _registry_error(
                "A registered source workbook could not be inspected",
                "registered_source_inaccessible",
                registry,
                row=row,
                registration_path=value,
                source_path=str(source),
                error_type=type(error).__name__,
            )
        if not stat.S_ISREG(source_identity.st_mode):
            _registry_error(
                "A registered source workbook is not a file",
                "registered_source_wrong_type",
                registry,
                row=row,
                registration_path=value,
                source_path=str(source),
            )
        for existing in result:
            if existing.identity is not None and os.path.samestat(
                source_identity,
                existing.identity,
            ):
                _registry_error(
                    "A workbook file is registered more than once",
                    "duplicate_registration_path",
                    registry,
                    row=row,
                    registration_path=value,
                    source_path=str(source),
                )
        result.append(
            _SourceSnapshot(
                kind="workbook",
                path=source,
                identity=source_identity,
                registry=registry,
                row=row,
                registration_path=value,
            )
        )
    return result


def _root_relative_posix(root: Path, path: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        _registry_error(
            "A model source resolves outside the project root",
            "unsafe_source_path",
            root / "Datas" / "__tables__.xlsx",
            source_path=str(path),
        )


def _stream_opened_source(
    digest: Any,
    source: _OpenedSource,
) -> None:
    try:
        source.handle.seek(0)
        while True:
            chunk = source.handle.read(_DIGEST_CHUNK_SIZE)
            if not chunk:
                break
            digest.update(chunk)
    except (OSError, MemoryError, ValueError) as error:
        _source_error(
            source.snapshot,
            "Unable to read an opened model source file",
            "source_read_error",
            error_type=type(error).__name__,
        )


def _registry_error(
    message: str,
    reason: str,
    registry: Path,
    **details: str | int,
) -> NoReturn:
    raise AuthoringError(
        "invalid_source_registry",
        message,
        {
            "path": str(registry),
            "reason": reason,
            "role": "source registry",
            **details,
        },
    )
