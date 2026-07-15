"""Marker-aware, loss-minimizing persistence for workbook-backed entities.

The adapter treats the Luban marker rows as the table schema, rather than
assuming that they occupy rows one through three.  Inspection results contain
only immutable primitives so authoring services can safely use them as a
snapshot of the current source.
"""

from __future__ import annotations

from contextlib import contextmanager
from copy import copy
from dataclasses import dataclass
from io import BytesIO
import os
from pathlib import Path
import re
import stat
import tempfile
from typing import Any, Iterator, NoReturn
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .change import ModelChange
from .entity_schema import (
    ENTITY_SCHEMAS,
    EntitySchema,
    FieldSpec,
    validate_entity_fields,
)
from .response import AuthoringError


_MARKERS = ("##var", "##", "##type")
_MAX_SOURCE_BYTES = 32 * 1024 * 1024
_MAX_ARCHIVE_BYTES = 64 * 1024 * 1024
_MAX_ARCHIVE_MEMBER_BYTES = 32 * 1024 * 1024
_MAX_ARCHIVE_MEMBERS = 2_048
_MAX_WORKSHEET_XML_BYTES = 16 * 1024 * 1024
_MAX_ROWS = 100_000
_MAX_COLUMNS = 1_024
_MAX_CELLS = 250_000
_CELL_REFERENCE_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
_STRING_CELL_TYPES = frozenset({"s", "inlineStr"})

_LOSSY_PACKAGE_PREFIXES = (
    ("_xmlsignatures/", "digital_signatures"),
    ("customui/", "custom_ui"),
    ("customxml/", "custom_xml"),
    ("xl/activex/", "activex_controls"),
    ("xl/connections", "data_connections"),
    ("xl/controls/", "worksheet_controls"),
    ("xl/ctrlprops/", "worksheet_controls"),
    ("xl/embeddings/", "embedded_objects"),
    ("xl/externallinks/", "external_links"),
    ("xl/metadata", "workbook_metadata"),
    ("xl/persons/", "threaded_comments"),
    ("xl/pivotcache/", "pivot_caches"),
    ("xl/printersettings/", "printer_settings"),
    ("xl/querytables/", "query_tables"),
    ("xl/revisions/", "shared_workbook_revisions"),
    ("xl/richdata/", "rich_data"),
    ("xl/slicercaches/", "slicers"),
    ("xl/slicers/", "slicers"),
    ("xl/threadedcomments/", "threaded_comments"),
    ("xl/timelinecaches/", "timelines"),
    ("xl/timelines/", "timelines"),
    ("xl/vbaproject", "vba_macros"),
    ("xl/webextensions/", "web_extensions"),
)


@dataclass(frozen=True, slots=True)
class _FileIdentity:
    device: int
    inode: int
    size: int
    modified_ns: int
    mode: int


@dataclass(frozen=True, slots=True)
class _WorkbookSnapshot:
    content: bytes
    identity: _FileIdentity
    source_mode: int


@dataclass(frozen=True, slots=True)
class WorkbookRecord:
    """One validated table row represented with immutable ordered fields."""

    entity_id: str
    row: int
    fields: tuple[tuple[str, Any], ...]


@dataclass(frozen=True, slots=True)
class TableInspection:
    """An immutable marker, column, and entity snapshot of one workbook."""

    path: Path
    entity: str
    sheet_name: str
    marker_column: int
    var_row: int
    description_row: int
    type_row: int
    columns: tuple[tuple[str, int], ...]
    records: tuple[WorkbookRecord, ...]
    duplicate_ids: tuple[str, ...]


def inspect_table(path: str | os.PathLike[str]) -> TableInspection:
    """Inspect and validate one canonical workbook without mutating it."""

    source = _coerce_path(path)
    schema = _schema_for_path(source)
    return _inspect_path(source, schema)


def find_duplicate_ids(path: str | os.PathLike[str]) -> list[str]:
    """Return duplicated entity ids once each, in first duplicate order."""

    return list(inspect_table(path).duplicate_ids)


def upsert_workbook_entity(
    path: str | os.PathLike[str],
    change: ModelChange,
) -> bool:
    """Create or update one complete workbook entity and replace atomically.

    The return value is ``False`` for a semantic no-op, in which case the
    workbook is never saved and its bytes remain untouched.
    """

    source = _coerce_path(path)
    schema = _schema_for_path(source)
    if change.entity != schema.entity:
        _source_error(
            "The change entity does not match the workbook",
            "entity_path_mismatch",
            change_entity=change.entity,
            entity=schema.entity,
            path=str(source),
        )

    snapshot = _read_snapshot(source)
    with _open_snapshot(snapshot, source, read_only=False) as workbook:
        inspected = _inspect_open(workbook, source, schema)
        if inspected.duplicate_ids:
            _source_error(
                "Workbook contains duplicate entity ids",
                "duplicate_ids",
                duplicate_ids=list(inspected.duplicate_ids),
                entity=schema.entity,
                path=str(source),
            )

        normalized = validate_entity_fields(
            schema.entity,
            change.id,
            _mutable_workbook_fields(schema, change.fields),
        )
        selected = [
            record for record in inspected.records if record.entity_id == change.id
        ]
        if selected and (
            _semantic_fields(schema, selected[0].fields)
            == _semantic_fields(schema, normalized)
        ):
            _require_current_identity(source, snapshot.identity)
            return False

        sheet = workbook.active
        if sheet is None:
            _source_error(
                "Workbook does not contain an active worksheet",
                "missing_active_sheet",
                path=str(source),
            )
        columns = dict(inspected.columns)
        if selected:
            target_row = selected[0].row
        else:
            target_row = max(sheet.max_row, inspected.type_row) + 1
            if target_row > _MAX_ROWS:
                _source_error(
                    "Appending an entity would exceed the workbook row limit",
                    "dimension_limit",
                    actual_rows=target_row,
                    limit_rows=_MAX_ROWS,
                    path=str(source),
                )
            style_row = (
                max(
                    (
                        record.row
                        for record in inspected.records
                        if record.row < target_row
                    ),
                    default=0,
                )
                or inspected.type_row
            )
            for field in ("id", *schema.field_names):
                column = columns[field]
                sheet.cell(target_row, column)._style = copy(
                    sheet.cell(style_row, column)._style
                )
            sheet.cell(target_row, columns["id"], change.id)

        current = dict(selected[0].fields) if selected else {}
        for field in schema.field_names:
            if selected and _semantic_value(field, current[field]) == _semantic_value(
                field, normalized[field]
            ):
                continue
            sheet.cell(
                target_row,
                columns[field],
                _encode_value(field, normalized[field]),
            )

        _replace_atomically(source, workbook, schema, change, snapshot)
        return True


def _coerce_path(path: str | os.PathLike[str]) -> Path:
    if not isinstance(path, (str, os.PathLike)):
        raise AuthoringError(
            "workbook_read_failed",
            "Workbook path must be path-like",
            {"reason": "invalid_path", "value_type": type(path).__name__},
        )
    return Path(path)


def _schema_for_path(path: Path) -> EntitySchema:
    for schema in ENTITY_SCHEMAS.values():
        if schema.storage_kind == "workbook" and schema.storage_name == path.name:
            return schema
    _source_error(
        "Workbook filename does not identify a supported authoring table",
        "unknown_table",
        path=str(path),
        supported=sorted(
            schema.storage_name
            for schema in ENTITY_SCHEMAS.values()
            if schema.storage_kind == "workbook"
        ),
    )


def _inspect_path(path: Path, schema: EntitySchema) -> TableInspection:
    snapshot = _read_snapshot(path)
    with _open_snapshot(snapshot, path, read_only=True) as workbook:
        return _inspect_open(workbook, path, schema)


def _read_snapshot(path: Path) -> _WorkbookSnapshot:
    try:
        with path.open("rb") as handle:
            before_stat = os.fstat(handle.fileno())
            before = _identity_from_stat(before_stat)
            if not stat.S_ISREG(before_stat.st_mode):
                raise OSError("workbook source is not a regular file")
            if before.size > _MAX_SOURCE_BYTES:
                _source_error(
                    "Workbook exceeds the compressed source size limit",
                    "source_too_large",
                    actual_bytes=before.size,
                    limit_bytes=_MAX_SOURCE_BYTES,
                    path=str(path),
                )
            content = handle.read(_MAX_SOURCE_BYTES + 1)
            after_stat = os.fstat(handle.fileno())
            after = _identity_from_stat(after_stat)
            source_mode = after.mode
    except AuthoringError:
        raise
    except (OSError, MemoryError, ValueError) as error:
        raise AuthoringError(
            "workbook_read_failed",
            "Workbook could not be read",
            {
                "error_type": type(error).__name__,
                "path": str(path),
                "reason": "read_error",
            },
        ) from None

    if len(content) > _MAX_SOURCE_BYTES:
        _source_error(
            "Workbook exceeds the compressed source size limit",
            "source_too_large",
            actual_bytes=len(content),
            limit_bytes=_MAX_SOURCE_BYTES,
            path=str(path),
        )
    if before != after or len(content) != after.size:
        _source_error(
            "Workbook changed while its immutable snapshot was being read",
            "source_changed_during_read",
            actual_bytes=len(content),
            expected_bytes=after.size,
            path=str(path),
        )

    snapshot = _WorkbookSnapshot(content, after, source_mode)
    _preflight_archive(snapshot, path)
    return snapshot


def _identity_from_stat(source_stat: os.stat_result) -> _FileIdentity:
    return _FileIdentity(
        device=source_stat.st_dev,
        inode=source_stat.st_ino,
        size=source_stat.st_size,
        modified_ns=source_stat.st_mtime_ns,
        mode=stat.S_IMODE(source_stat.st_mode),
    )


def _preflight_archive(snapshot: _WorkbookSnapshot, path: Path) -> None:
    try:
        with ZipFile(BytesIO(snapshot.content)) as archive:
            members = archive.infolist()
            if len(members) > _MAX_ARCHIVE_MEMBERS:
                _source_error(
                    "Workbook archive contains too many members",
                    "archive_limit",
                    actual_members=len(members),
                    limit_members=_MAX_ARCHIVE_MEMBERS,
                    path=str(path),
                )
            total_size = 0
            names: set[str] = set()
            worksheets = []
            for member in members:
                if member.filename in names:
                    _source_error(
                        "Workbook archive contains a duplicate package part",
                        "unsafe_archive",
                        member=member.filename,
                        path=str(path),
                    )
                names.add(member.filename)
                if member.flag_bits & 0x1:
                    _source_error(
                        "Encrypted workbook members are not supported",
                        "unsafe_archive",
                        member=member.filename,
                        path=str(path),
                    )
                if member.file_size > _MAX_ARCHIVE_MEMBER_BYTES:
                    _source_error(
                        "Workbook archive member exceeds the size limit",
                        "archive_limit",
                        actual_bytes=member.file_size,
                        limit_bytes=_MAX_ARCHIVE_MEMBER_BYTES,
                        member=member.filename,
                        path=str(path),
                    )
                total_size += member.file_size
                if total_size > _MAX_ARCHIVE_BYTES:
                    _source_error(
                        "Workbook expanded archive exceeds the size limit",
                        "archive_limit",
                        actual_bytes=total_size,
                        limit_bytes=_MAX_ARCHIVE_BYTES,
                        path=str(path),
                    )
                lowered = member.filename.lower()
                for prefix, feature in _LOSSY_PACKAGE_PREFIXES:
                    if lowered.startswith(prefix):
                        _source_error(
                            "Workbook contains a package part that cannot be preserved safely",
                            "unsupported_package_part",
                            feature=feature,
                            part=member.filename,
                            path=str(path),
                        )
                if lowered.startswith("xl/worksheets/") and lowered.endswith(".xml"):
                    worksheets.append(member)

            for worksheet in worksheets:
                _preflight_worksheet(archive, worksheet, path)
    except AuthoringError:
        raise
    except (BadZipFile, OSError, ValueError, RuntimeError) as error:
        raise AuthoringError(
            "invalid_workbook_source",
            "Workbook is not a readable OOXML archive",
            {
                "error_type": type(error).__name__,
                "path": str(path),
                "reason": "corrupt_workbook",
            },
        ) from None


class _BudgetedWorksheetReader:
    def __init__(self, source: Any, path: Path, part: str) -> None:
        self._source = source
        self._path = path
        self._part = part
        self.bytes_read = 0

    def read(self, size: int = -1) -> bytes:
        data = self._source.read(size)
        self.bytes_read += len(data)
        if self.bytes_read > _MAX_WORKSHEET_XML_BYTES:
            _source_error(
                "Worksheet XML exceeds the streaming size limit",
                "worksheet_xml_limit",
                actual_bytes=self.bytes_read,
                limit_bytes=_MAX_WORKSHEET_XML_BYTES,
                part=self._part,
                path=str(self._path),
            )
        return data


def _preflight_worksheet(archive: ZipFile, member: Any, path: Path) -> None:
    if member.file_size > _MAX_WORKSHEET_XML_BYTES:
        _source_error(
            "Worksheet XML exceeds the streaming size limit",
            "worksheet_xml_limit",
            actual_bytes=member.file_size,
            limit_bytes=_MAX_WORKSHEET_XML_BYTES,
            part=member.filename,
            path=str(path),
        )

    cell_count = 0
    max_row = 0
    max_column = 0
    try:
        with archive.open(member) as raw:
            reader = _BudgetedWorksheetReader(raw, path, member.filename)
            for event, element in ElementTree.iterparse(
                reader,
                events=("start", "end"),
            ):
                if event == "start":
                    local_name = element.tag.rsplit("}", 1)[-1]
                    if local_name == "dimension":
                        reference = element.attrib.get("ref")
                        if not reference:
                            _invalid_worksheet_xml(path, member.filename, "missing_dimension_ref")
                        dimension_parts = reference.split(":")
                        if len(dimension_parts) not in {1, 2}:
                            _invalid_worksheet_xml(path, member.filename, "invalid_dimension_ref")
                        first_column, first_row = _parse_cell_reference(
                            dimension_parts[0], path, member.filename
                        )
                        last_column, last_row = _parse_cell_reference(
                            dimension_parts[-1], path, member.filename
                        )
                        if last_column < first_column or last_row < first_row:
                            _invalid_worksheet_xml(path, member.filename, "reversed_dimension_ref")
                        _enforce_worksheet_budget(
                            last_row,
                            last_column,
                            cell_count,
                            path,
                            member.filename,
                        )
                    elif local_name == "row":
                        row_text = element.attrib.get("r")
                        if row_text is not None:
                            if not row_text.isdigit() or int(row_text) < 1:
                                _invalid_worksheet_xml(path, member.filename, "invalid_row_number")
                            max_row = max(max_row, int(row_text))
                            _enforce_worksheet_budget(
                                max_row,
                                max_column,
                                cell_count,
                                path,
                                member.filename,
                            )
                    elif local_name == "c":
                        reference = element.attrib.get("r")
                        if not reference:
                            _invalid_worksheet_xml(path, member.filename, "missing_cell_ref")
                        column, row = _parse_cell_reference(reference, path, member.filename)
                        cell_count += 1
                        max_row = max(max_row, row)
                        max_column = max(max_column, column)
                        _enforce_worksheet_budget(
                            max_row,
                            max_column,
                            cell_count,
                            path,
                            member.filename,
                        )
                else:
                    element.clear()
    except AuthoringError:
        raise
    except (BadZipFile, ElementTree.ParseError, OSError, RuntimeError, ValueError) as error:
        raise AuthoringError(
            "invalid_workbook_source",
            "Worksheet XML could not be streamed safely",
            {
                "error_type": type(error).__name__,
                "part": member.filename,
                "path": str(path),
                "reason": "corrupt_workbook",
            },
        ) from None


def _parse_cell_reference(reference: str, path: Path, part: str) -> tuple[int, int]:
    matched = _CELL_REFERENCE_RE.fullmatch(reference)
    if matched is None:
        _invalid_worksheet_xml(path, part, "invalid_cell_ref")
    column = 0
    for character in matched.group(1):
        column = column * 26 + ord(character) - ord("A") + 1
    return column, int(matched.group(2))


def _enforce_worksheet_budget(
    rows: int,
    columns: int,
    cells: int,
    path: Path,
    part: str,
) -> None:
    if (
        rows > _MAX_ROWS
        or columns > _MAX_COLUMNS
        or rows * columns > _MAX_CELLS
        or cells > _MAX_CELLS
    ):
        _source_error(
            "Workbook dimensions exceed the authoring safety limit",
            "dimension_limit",
            actual_cells=cells,
            actual_columns=columns,
            actual_rows=rows,
            limit_cells=_MAX_CELLS,
            limit_columns=_MAX_COLUMNS,
            limit_rows=_MAX_ROWS,
            part=part,
            path=str(path),
        )


def _invalid_worksheet_xml(path: Path, part: str, detail: str) -> NoReturn:
    _source_error(
        "Worksheet XML contains an invalid coordinate or dimension",
        "corrupt_workbook",
        detail=detail,
        part=part,
        path=str(path),
    )


@contextmanager
def _open_snapshot(
    snapshot: _WorkbookSnapshot,
    path: Path,
    *,
    read_only: bool,
) -> Iterator[Workbook]:
    buffer = BytesIO(snapshot.content)
    workbook: Workbook | None = None
    try:
        try:
            workbook = load_workbook(
                buffer,
                read_only=read_only,
                data_only=False,
                keep_links=False,
                rich_text=True,
            )
        except Exception as error:
            raise AuthoringError(
                "invalid_workbook_source",
                "Workbook could not be opened safely",
                {
                    "error_type": type(error).__name__,
                    "path": str(path),
                    "reason": "corrupt_workbook",
                },
            ) from None
        yield workbook
    finally:
        if workbook is not None:
            workbook.close()
        buffer.close()


def _inspect_open(
    workbook: Workbook,
    path: Path,
    schema: EntitySchema,
) -> TableInspection:
    sheet = workbook.active
    if sheet is None:
        _source_error(
            "Workbook does not contain an active worksheet",
            "missing_active_sheet",
            path=str(path),
        )
    _validate_dimensions(sheet, path)
    marker_positions: dict[str, list[tuple[int, int]]] = {
        marker: [] for marker in _MARKERS
    }
    for row in sheet.iter_rows():
        for cell in row:
            if type(cell.value) is str and cell.value in marker_positions:
                marker_positions[cell.value].append((cell.row, cell.column))

    for marker in _MARKERS:
        positions = marker_positions[marker]
        if not positions:
            _source_error(
                "Workbook is missing a Luban marker",
                "missing_marker",
                marker=marker,
                path=str(path),
            )
        if len(positions) != 1:
            _source_error(
                "Workbook contains a duplicate Luban marker",
                "duplicate_marker",
                marker=marker,
                path=str(path),
                positions=[list(position) for position in positions[:16]],
            )

    var_row, marker_column = marker_positions["##var"][0]
    description_row, description_column = marker_positions["##"][0]
    type_row, type_column = marker_positions["##type"][0]
    if not (
        marker_column == description_column == type_column
        and var_row < description_row < type_row
    ):
        _source_error(
            "Workbook marker rows do not form one coherent table schema",
            "incoherent_markers",
            marker_positions=[
                [marker, *marker_positions[marker][0]] for marker in _MARKERS
            ],
            path=str(path),
        )

    header_columns: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        if column == marker_column:
            continue
        raw_header = sheet.cell(var_row, column).value
        if raw_header in (None, ""):
            continue
        header = str(raw_header)
        if header in header_columns:
            _source_error(
                "Workbook contains a duplicate header",
                "duplicate_header",
                column=column,
                first_column=header_columns[header],
                header=header,
                path=str(path),
            )
        header_columns[header] = column

    expected_fields = ("id", *schema.field_names)
    for field in expected_fields:
        if field not in header_columns:
            _source_error(
                "Workbook is missing a required schema column",
                "missing_column",
                column=field,
                entity=schema.entity,
                path=str(path),
            )
        column = header_columns[field]
        actual_type = sheet.cell(type_row, column).value
        expected_type = _expected_column_type(schema.entity, field)
        if actual_type != expected_type:
            _source_error(
                "Workbook schema column has the wrong Luban type",
                "wrong_column_type",
                actual=actual_type if isinstance(actual_type, (str, int, bool)) else None,
                column=field,
                expected=expected_type,
                path=str(path),
            )

    columns = tuple((field, header_columns[field]) for field in expected_fields)
    records: list[WorkbookRecord] = []
    duplicates: list[str] = []
    reported: set[str] = set()
    seen: set[str] = set()
    id_column = header_columns["id"]
    field_columns = {field: header_columns[field] for field in schema.field_names}
    field_specs = {field.name: field for field in schema.fields}
    for row_index in range(type_row + 1, sheet.max_row + 1):
        id_cell = sheet.cell(row_index, id_column)
        raw_id = id_cell.value
        field_cells = {
            field: sheet.cell(row_index, column)
            for field, column in field_columns.items()
        }
        if raw_id in (None, ""):
            if any(cell.value not in (None, "") for cell in field_cells.values()):
                _source_error(
                    "Workbook row has schema values but no entity id",
                    "missing_id",
                    entity=schema.entity,
                    path=str(path),
                    row=row_index,
                )
            continue
        try:
            entity_id = _decode_id_cell(schema.entity, id_cell)
            decoded = {
                field: _decode_field_cell(
                    schema.entity,
                    entity_id,
                    field_specs[field],
                    field_cells[field],
                )
                for field in schema.field_names
            }
            normalized = validate_entity_fields(schema.entity, entity_id, decoded)
        except AuthoringError as error:
            _source_error(
                "Workbook contains an invalid entity row",
                "invalid_entity_row",
                allowed=error.details.get("allowed", ()),
                entity=schema.entity,
                field=error.details.get("field"),
                id=error.details.get("id", _diagnostic_value(raw_id)),
                path=str(path),
                row=row_index,
                source_type=error.details.get("source_type"),
                validation_code=error.code,
                validation_message=error.message,
                value=error.details.get("value"),
            )
        record_fields = tuple(
            (field, _immutable_value(field, normalized[field]))
            for field in schema.field_names
        )
        records.append(WorkbookRecord(entity_id, row_index, record_fields))
        if entity_id in seen and entity_id not in reported:
            duplicates.append(entity_id)
            reported.add(entity_id)
        seen.add(entity_id)

    return TableInspection(
        path=path,
        entity=schema.entity,
        sheet_name=sheet.title,
        marker_column=marker_column,
        var_row=var_row,
        description_row=description_row,
        type_row=type_row,
        columns=columns,
        records=tuple(records),
        duplicate_ids=tuple(duplicates),
    )


def _validate_dimensions(sheet: Worksheet, path: Path) -> None:
    rows = sheet.max_row
    columns = sheet.max_column
    if rows > _MAX_ROWS or columns > _MAX_COLUMNS or rows * columns > _MAX_CELLS:
        _source_error(
            "Workbook dimensions exceed the authoring safety limit",
            "dimension_limit",
            actual_columns=columns,
            actual_rows=rows,
            limit_cells=_MAX_CELLS,
            limit_columns=_MAX_COLUMNS,
            limit_rows=_MAX_ROWS,
            path=str(path),
        )


def _expected_column_type(entity: str, field: str) -> str:
    if entity == "prestige_layer" and field == "reset_resources":
        return "(list#sep=;),string"
    return "string"


def _decode_id_cell(entity: str, cell: Any) -> str:
    value = cell.value
    if cell.data_type not in _STRING_CELL_TYPES or type(value) is not str:
        _invalid_cell_value(
            entity,
            _diagnostic_value(value),
            "id",
            value,
            ("literal string id",),
            cell.data_type,
        )
    return value


def _decode_field_cell(
    entity: str,
    entity_id: str,
    spec: FieldSpec,
    cell: Any,
) -> Any:
    value = cell.value
    if spec.value_type == "list_id":
        if value in (None, ""):
            return []
        if cell.data_type not in _STRING_CELL_TYPES or type(value) is not str:
            _invalid_cell_value(
                entity,
                entity_id,
                spec.name,
                value,
                ("semicolon-separated string ids",),
                cell.data_type,
            )
        values = value.split(";")
        if any(item == "" for item in values):
            _invalid_cell_value(
                entity,
                entity_id,
                spec.name,
                value,
                ("semicolon-separated string ids without empty items",),
                cell.data_type,
            )
        return values

    if spec.value_type in {"decimal", "positive_decimal", "nonnegative_decimal"}:
        is_integer_cell = cell.data_type == "n" and type(value) is int
        is_string_cell = (
            cell.data_type in _STRING_CELL_TYPES and type(value) is str
        )
        if not is_integer_cell and not is_string_cell:
            _invalid_cell_value(
                entity,
                entity_id,
                spec.name,
                value,
                ("integer cell", "exact decimal string cell"),
                cell.data_type,
            )
        return value

    if cell.data_type not in _STRING_CELL_TYPES or type(value) is not str:
        _invalid_cell_value(
            entity,
            entity_id,
            spec.name,
            value,
            ("literal string cell",),
            cell.data_type,
        )
    return value


def _invalid_cell_value(
    entity: str,
    entity_id: Any,
    field: str,
    value: Any,
    allowed: tuple[str, ...],
    source_type: str,
) -> NoReturn:
    raise AuthoringError(
        "invalid_change",
        f"Workbook cell {field!r} has an unsupported exact value type",
        {
            "allowed": allowed,
            "entity": entity,
            "field": field,
            "id": entity_id,
            "source_type": source_type,
            "value": _diagnostic_value(value),
            "value_type": type(value).__name__,
        },
    )


def _diagnostic_value(value: Any) -> str | int | bool | None:
    if value is None or type(value) in {str, int, bool}:
        return value
    return f"<{type(value).__name__}>"


def _encode_value(field: str, value: Any) -> str:
    if field == "reset_resources":
        return ";".join(value)
    return str(value)


def _immutable_value(field: str, value: Any) -> Any:
    if field == "reset_resources":
        return tuple(value)
    return value


def _mutable_workbook_fields(
    schema: EntitySchema,
    fields: Any,
) -> dict[str, Any]:
    """Thaw the one list-valued workbook field frozen by ``ModelChange``."""

    result = dict(fields)
    if schema.entity == "prestige_layer" and "reset_resources" in result:
        result["reset_resources"] = list(result["reset_resources"])
    return result


def _semantic_value(field: str, value: Any) -> Any:
    if field == "reset_resources":
        return tuple(value)
    return str(value)


def _semantic_fields(
    schema: EntitySchema,
    fields: Any,
) -> tuple[tuple[str, Any], ...]:
    """Canonicalize field equality in schema order, never mapping order."""

    values = dict(fields)
    return tuple(
        (field, _semantic_value(field, values[field]))
        for field in schema.field_names
    )


def _replace_atomically(
    path: Path,
    workbook: Workbook,
    schema: EntitySchema,
    change: ModelChange,
    source_snapshot: _WorkbookSnapshot,
) -> None:
    temporary: Path | None = None
    phase = "temp_write"
    candidate_bytes: bytes | None = None
    try:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{path.name}.",
            suffix=".tmp.xlsx",
            dir=path.parent,
        )
        os.close(descriptor)
        temporary = Path(temporary_name)
        workbook.save(temporary)
        os.chmod(temporary, source_snapshot.source_mode)
        candidate_bytes = temporary.read_bytes()

        phase = "reload"
        reloaded = _inspect_path(temporary, schema)
        matches = [record for record in reloaded.records if record.entity_id == change.id]
        if len(matches) != 1 or (
            _semantic_fields(schema, matches[0].fields)
            != _semantic_fields(schema, change.fields)
        ):
            _source_error(
                "Reloaded workbook does not contain the exact upserted entity",
                "reload_mismatch",
                entity=change.entity,
                id=change.id,
                path=str(path),
            )

        phase = "identity"
        _require_current_identity(path, source_snapshot.identity)
        phase = "replace"
        try:
            os.replace(temporary, path)
        except BaseException:
            if candidate_bytes is not None and _path_has_exact_bytes(path, candidate_bytes):
                return
            raise
        temporary = None
    except Exception as error:
        reason = {
            "temp_write": "temp_write_error",
            "reload": "reload_error",
            "identity": "source_changed",
            "replace": "replace_error",
        }[phase]
        raise AuthoringError(
            "workbook_write_failed",
            "Workbook candidate could not be replaced atomically",
            {
                "error_type": type(error).__name__,
                "path": str(path),
                "reason": reason,
            },
        ) from None
    finally:
        if temporary is not None:
            try:
                temporary.unlink(missing_ok=True)
            except OSError:
                pass


def _path_has_exact_bytes(path: Path, expected: bytes) -> bool:
    try:
        return path.stat().st_size == len(expected) and path.read_bytes() == expected
    except (OSError, MemoryError, ValueError):
        return False


def _require_current_identity(path: Path, expected: _FileIdentity) -> None:
    try:
        actual = _identity_from_stat(path.stat())
    except (OSError, ValueError):
        actual = None
    if actual != expected:
        _source_error(
            "Workbook path changed after its source snapshot was opened",
            "source_changed",
            path=str(path),
        )


def _source_error(message: str, reason: str, **details: Any) -> NoReturn:
    raise AuthoringError(
        "invalid_workbook_source",
        message,
        {"reason": reason, **details},
    )


__all__ = [
    "TableInspection",
    "WorkbookRecord",
    "find_duplicate_ids",
    "inspect_table",
    "upsert_workbook_entity",
]
