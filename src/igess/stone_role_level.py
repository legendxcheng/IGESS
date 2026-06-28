from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .numbers import SimNumber


@dataclass(frozen=True)
class AttributeDefinition:
    key: str
    value_type: str
    power_value: SimNumber
    enabled: bool


@dataclass(frozen=True)
class RoleLevelRow:
    level: int
    exp_req: SimNumber
    cumulative_exp_to_level_start: SimNumber
    cumulative_exp_to_next_level: SimNumber
    combat_power: SimNumber
    combat_power_delta: SimNumber | None


@dataclass(frozen=True)
class RoleLevelCurve:
    role_lv_path: Path
    attribute_def_path: Path
    rows: list[RoleLevelRow]


@dataclass(frozen=True)
class RealmProgressionRow:
    realm_id: int
    realm_name: str
    level_cap: int
    realm_combat_power: SimNumber
    realm_combat_power_delta: SimNumber | None


@dataclass(frozen=True)
class RealmProgressionCurve:
    role_realm_path: Path
    attribute_def_path: Path
    rows: list[RealmProgressionRow]


@dataclass(frozen=True)
class _ColumnSpec:
    key: str
    value_type: str
    indexes: tuple[int, ...]


def build_role_level_curve(
    role_lv_path: str | Path,
    attribute_def_path: str | Path,
) -> RoleLevelCurve:
    role_lv = Path(role_lv_path)
    attribute_def = Path(attribute_def_path)
    definitions = _load_attribute_definitions(attribute_def)
    role_rows = _load_role_level_rows(role_lv)
    curve_rows: list[RoleLevelRow] = []
    cumulative_exp = SimNumber.zero()
    previous_power: SimNumber | None = None
    for values in role_rows:
        level = int(values["id"].decimal)
        exp_req = values["expReq"]
        combat_power = _calculate_combat_power(values, definitions)
        curve_rows.append(
            RoleLevelRow(
                level=level,
                exp_req=exp_req,
                cumulative_exp_to_level_start=cumulative_exp,
                cumulative_exp_to_next_level=cumulative_exp + exp_req,
                combat_power=combat_power,
                combat_power_delta=None
                if previous_power is None
                else combat_power - previous_power,
            )
        )
        cumulative_exp += exp_req
        previous_power = combat_power
    return RoleLevelCurve(
        role_lv_path=role_lv,
        attribute_def_path=attribute_def,
        rows=curve_rows,
    )


def build_realm_progression_curve(
    role_realm_path: str | Path,
    attribute_def_path: str | Path,
) -> RealmProgressionCurve:
    role_realm = Path(role_realm_path)
    attribute_def = Path(attribute_def_path)
    definitions = _load_attribute_definitions(attribute_def)
    realm_rows = _load_role_realm_rows(role_realm, definitions)
    curve_rows: list[RealmProgressionRow] = []
    previous_power: SimNumber | None = None
    for realm_id, realm_name, level_cap, values in realm_rows:
        combat_power = _calculate_combat_power(values, definitions)
        curve_rows.append(
            RealmProgressionRow(
                realm_id=realm_id,
                realm_name=realm_name,
                level_cap=level_cap,
                realm_combat_power=combat_power,
                realm_combat_power_delta=None
                if previous_power is None
                else combat_power - previous_power,
            )
        )
        previous_power = combat_power
    return RealmProgressionCurve(
        role_realm_path=role_realm,
        attribute_def_path=attribute_def,
        rows=curve_rows,
    )


def write_role_level_artifacts(result: RoleLevelCurve, output_dir: str | Path) -> dict[str, Path]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    curve_json = output_path / "role_level_curve.json"
    curve_csv = output_path / "role_level_curve.csv"
    summary_md = output_path / "role_level_summary.md"
    manifest_json = output_path / "source_manifest.json"

    payload = [_row_payload(row) for row in result.rows]
    curve_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )

    with curve_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "level",
                "exp_req",
                "cumulative_exp_to_level_start",
                "cumulative_exp_to_next_level",
                "combat_power",
                "combat_power_delta",
            ],
        )
        writer.writeheader()
        writer.writerows(payload)

    summary_md.write_text(_summary_markdown(result), encoding="utf-8", newline="\n")
    manifest_json.write_text(
        json.dumps(_source_manifest(result), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return {
        "curve_json": curve_json,
        "curve_csv": curve_csv,
        "summary_md": summary_md,
        "manifest_json": manifest_json,
    }


def write_realm_progression_artifacts(
    result: RealmProgressionCurve,
    output_dir: str | Path,
) -> dict[str, Path]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    curve_json = output_path / "realm_progression_curve.json"
    curve_csv = output_path / "realm_progression_curve.csv"
    summary_md = output_path / "realm_progression_summary.md"
    manifest_json = output_path / "source_manifest.json"

    payload = [_realm_row_payload(row) for row in result.rows]
    curve_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )

    with curve_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "realm_id",
                "realm_name",
                "level_cap",
                "realm_combat_power",
                "realm_combat_power_delta",
            ],
        )
        writer.writeheader()
        writer.writerows(payload)

    summary_md.write_text(_realm_summary_markdown(result), encoding="utf-8", newline="\n")
    manifest_json.write_text(
        json.dumps(_realm_source_manifest(result), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return {
        "curve_json": curve_json,
        "curve_csv": curve_csv,
        "summary_md": summary_md,
        "manifest_json": manifest_json,
    }


def _load_attribute_definitions(path: Path) -> dict[str, AttributeDefinition]:
    rows = _read_workbook_rows(path)
    if len(rows) < 4:
        raise ValueError(f"{path} does not contain attribute definition rows")
    headers = [str(value) if value is not None else "" for value in rows[0]]
    definitions: dict[str, AttributeDefinition] = {}
    for row in rows[3:]:
        data = dict(zip(headers, row))
        key = data.get("key")
        if key in (None, ""):
            continue
        definitions[str(key)] = AttributeDefinition(
            key=str(key),
            value_type=str(data.get("valueType") or ""),
            power_value=_sim_number(data.get("powerValue")),
            enabled=int(data.get("enabled") or 0) == 1,
        )
    return definitions


def _load_role_level_rows(path: Path) -> list[dict[str, SimNumber]]:
    rows = _read_workbook_rows(path)
    if len(rows) < 6:
        raise ValueError(f"{path} does not contain RoleLv data rows")
    column_specs = _role_level_column_specs(rows[0], rows[2])
    parsed_rows: list[dict[str, SimNumber]] = []
    for row in rows[5:]:
        if all(value in (None, "") for value in row[1:]):
            continue
        values: dict[str, SimNumber] = {}
        for spec in column_specs:
            if spec.value_type == "BigNumberParts":
                sign_index, coeff_index, exp_index = spec.indexes
                values[spec.key] = _big_number_parts(
                    row[sign_index],
                    row[coeff_index],
                    row[exp_index],
                )
            else:
                values[spec.key] = _sim_number(row[spec.indexes[0]])
        parsed_rows.append(values)
    return parsed_rows


def _load_role_realm_rows(
    path: Path,
    definitions: dict[str, AttributeDefinition],
) -> list[tuple[int, str, int, dict[str, SimNumber]]]:
    rows = _read_workbook_rows(path)
    if len(rows) < 5:
        raise ValueError(f"{path} does not contain RoleRealm data rows")
    column_specs = _role_level_column_specs(rows[0], rows[1])
    parsed_rows: list[tuple[int, str, int, dict[str, SimNumber]]] = []
    for row in rows[4:]:
        if all(value in (None, "") for value in row[1:]):
            continue
        realm_id = 0
        realm_name = ""
        level_cap = 0
        values: dict[str, SimNumber] = {}
        for spec in column_specs:
            if spec.key == "id":
                realm_id = int(_sim_number(row[spec.indexes[0]]).decimal)
                continue
            if spec.key == "name":
                realm_name = str(row[spec.indexes[0]] or "")
                continue
            if spec.key == "lvl_up":
                level_cap = int(_sim_number(row[spec.indexes[0]]).decimal)
                continue
            if spec.key not in definitions:
                continue
            if spec.value_type == "BigNumberParts":
                sign_index, coeff_index, exp_index = spec.indexes
                values[spec.key] = _big_number_parts(
                    row[sign_index],
                    row[coeff_index],
                    row[exp_index],
                )
            else:
                values[spec.key] = _sim_number(row[spec.indexes[0]])
        parsed_rows.append((realm_id, realm_name, level_cap, values))
    return parsed_rows


def _role_level_column_specs(header_row: tuple[Any, ...], type_row: tuple[Any, ...]) -> list[_ColumnSpec]:
    specs: list[_ColumnSpec] = []
    index = 1
    while index < len(header_row):
        key = header_row[index]
        field_type = type_row[index]
        if key in (None, ""):
            index += 1
            continue
        if field_type == "BigNumberParts":
            specs.append(
                _ColumnSpec(
                    key=str(key),
                    value_type="BigNumberParts",
                    indexes=(index, index + 1, index + 2),
                )
            )
            index += 3
        else:
            specs.append(
                _ColumnSpec(
                    key=str(key),
                    value_type=str(field_type or ""),
                    indexes=(index,),
                )
            )
            index += 1
    return specs


def _calculate_combat_power(
    values: dict[str, SimNumber],
    definitions: dict[str, AttributeDefinition],
) -> SimNumber:
    power = SimNumber.zero()
    for key, value in values.items():
        definition = definitions.get(key)
        if (
            definition is None
            or not definition.enabled
            or definition.power_value <= SimNumber.zero()
        ):
            continue
        normalized = (
            value / SimNumber.parse(10000)
            if definition.value_type == "ratio_bps"
            else value
        )
        power += normalized * definition.power_value
    return power


def _row_payload(row: RoleLevelRow) -> dict[str, int | str | None]:
    return {
        "level": row.level,
        "exp_req": _format_decimal(row.exp_req),
        "cumulative_exp_to_level_start": _format_decimal(row.cumulative_exp_to_level_start),
        "cumulative_exp_to_next_level": _format_decimal(row.cumulative_exp_to_next_level),
        "combat_power": _format_decimal(row.combat_power),
        "combat_power_delta": None
        if row.combat_power_delta is None
        else _format_decimal(row.combat_power_delta),
    }


def _realm_row_payload(row: RealmProgressionRow) -> dict[str, int | str | None]:
    return {
        "realm_id": row.realm_id,
        "realm_name": row.realm_name,
        "level_cap": row.level_cap,
        "realm_combat_power": _format_decimal(row.realm_combat_power),
        "realm_combat_power_delta": None
        if row.realm_combat_power_delta is None
        else _format_decimal(row.realm_combat_power_delta),
    }


def _summary_markdown(result: RoleLevelCurve) -> str:
    level_count = len(result.rows)
    first = result.rows[0] if result.rows else None
    last = result.rows[-1] if result.rows else None
    lines = [
        "# Stone Role Level Baseline",
        "",
        f"RoleLv source: `{result.role_lv_path}`",
        f"Attribute definition source: `{result.attribute_def_path}`",
        "",
        "Number backend: `bignum_log` (`igess.numbers.SimNumber`)",
        f"Level count: {level_count}",
    ]
    if first is not None and last is not None:
        lines.extend(
            [
                f"Min level: {first.level}",
                f"Max level: {last.level}",
                f"Level 1 combat power: {_format_decimal(first.combat_power)}",
                f"Level {last.level} combat power: {_format_decimal(last.combat_power)}",
                "Cumulative exp to max level start: "
                f"{_format_decimal(last.cumulative_exp_to_level_start)}",
            ]
        )
    lines.extend(
        [
            "",
            "Formula:",
            "",
            "- `BigNumberParts = sign * coeff * 10^exp`",
            "- `big_number/integer contribution = value * powerValue`",
            "- `ratio_bps contribution = value / 10000 * powerValue`",
            "- `combat_power = sum(contributions)`",
            "",
        ]
    )
    return "\n".join(lines)


def _realm_summary_markdown(result: RealmProgressionCurve) -> str:
    realm_count = len(result.rows)
    first = result.rows[0] if result.rows else None
    last = result.rows[-1] if result.rows else None
    lines = [
        "# Stone Realm Progression Baseline",
        "",
        f"RoleRealm source: `{result.role_realm_path}`",
        f"Attribute definition source: `{result.attribute_def_path}`",
        "",
        "Number backend: `bignum_log` (`igess.numbers.SimNumber`)",
        f"Realm count: {realm_count}",
        "Level combat power is not included; `level_cap` is metadata only.",
    ]
    if first is not None and last is not None:
        lines.extend(
            [
                f"First realm: {first.realm_id} {first.realm_name}",
                f"Last realm: {last.realm_id} {last.realm_name}",
                "First realm combat power: "
                f"{_format_decimal(first.realm_combat_power)}",
                "Last realm combat power: "
                f"{_format_decimal(last.realm_combat_power)}",
            ]
        )
    lines.extend(
        [
            "",
            "Formula:",
            "",
            "- `BigNumberParts = sign * coeff * 10^exp`",
            "- `big_number/integer contribution = value * powerValue`",
            "- `ratio_bps contribution = value / 10000 * powerValue`",
            "- `realm_combat_power = sum(contributions)`",
            "",
        ]
    )
    return "\n".join(lines)


def _source_manifest(result: RoleLevelCurve) -> dict[str, Any]:
    return {
        "project": "stone",
        "model": "role_level_baseline",
        "number_backend": "bignum_log",
        "number_type": "igess.numbers.SimNumber",
        "sources": {
            "role_lv": str(result.role_lv_path),
            "attribute_def": str(result.attribute_def_path),
        },
        "artifacts": [
            "role_level_curve.json",
            "role_level_curve.csv",
            "role_level_summary.md",
            "source_manifest.json",
        ],
        "formula": {
            "big_number_parts": "sign * coeff * 10^exp",
            "big_number_or_integer": "value * powerValue",
            "ratio_bps": "value / 10000 * powerValue",
        },
        "level_count": len(result.rows),
        "max_level": result.rows[-1].level if result.rows else None,
    }


def _realm_source_manifest(result: RealmProgressionCurve) -> dict[str, Any]:
    return {
        "project": "stone",
        "model": "realm_progression_baseline",
        "number_backend": "bignum_log",
        "number_type": "igess.numbers.SimNumber",
        "sources": {
            "role_realm": str(result.role_realm_path),
            "attribute_def": str(result.attribute_def_path),
        },
        "artifacts": [
            "realm_progression_curve.json",
            "realm_progression_curve.csv",
            "realm_progression_summary.md",
            "source_manifest.json",
        ],
        "formula": {
            "big_number_parts": "sign * coeff * 10^exp",
            "big_number_or_integer": "value * powerValue",
            "ratio_bps": "value / 10000 * powerValue",
            "level_combat_power": "not included",
        },
        "realm_count": len(result.rows),
        "max_realm_id": result.rows[-1].realm_id if result.rows else None,
    }


def _big_number_parts(sign: Any, coeff: Any, exponent: Any) -> SimNumber:
    return _sim_number(sign) * _sim_number(coeff) * (
        SimNumber.parse(10) ** int(_sim_number(exponent).decimal)
    )


def _sim_number(value: Any) -> SimNumber:
    if value in (None, ""):
        return SimNumber.zero()
    return SimNumber.parse(value)


def _format_decimal(value: SimNumber) -> str:
    return value.to_decimal_string()


def _read_workbook_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))
