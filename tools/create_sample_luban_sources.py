from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from igess.luban_exporter import export_registered_workbooks


ROOT = Path(__file__).resolve().parents[1]
DATAS = ROOT / "data-tables" / "Datas"
EXPORTS = ROOT / "examples" / "shelldiver_v0" / "luban_exports"


def write_workbook(path: Path, headers: list[str], comments: list[str], types: list[str], rows: list[list]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = path.stem
    sheet.append(["##var", *headers])
    sheet.append(["##", *comments])
    sheet.append(["##type", *types])
    for row in rows:
        sheet.append(["", *row])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor="F7F1D9")
    for cell in sheet[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    write_workbook(
        DATAS / "__tables__.xlsx",
        ["table", "path", "mode", "key"],
        ["stable table id", "source workbook", "export mode", "map key field"],
        ["string", "string", "string", "string"],
        [
            ["resources", "resources.xlsx", "map", "id"],
            ["generators", "generators.xlsx", "map", "id"],
            ["upgrades", "upgrades.xlsx", "map", "id"],
            ["constants", "constants.xlsx", "map", "id"],
            ["milestones", "milestones.xlsx", "map", "id"],
            ["prestige_layers", "prestige_layers.xlsx", "map", "id"],
        ],
    )
    write_workbook(
        DATAS / "resources.xlsx",
        ["id", "name", "dimension"],
        ["stable resource id", "display name", "quantity dimension"],
        ["string", "string", "string"],
        [["fish", "Fish", "fish"], ["prestige_point", "Prestige Point", "prestige"]],
    )
    write_workbook(
        DATAS / "generators.xlsx",
        [
            "id",
            "name",
            "generator_type",
            "output_resource",
            "source_type",
            "base_output",
            "base_cost",
            "cost_resource",
            "cost_growth",
            "unlock_condition",
        ],
        [
            "stable generator id",
            "display name",
            "YAML generator type",
            "produced resource id",
            "source type id",
            "base output per second",
            "first purchase cost",
            "resource spent",
            "exponential cost growth",
            "deterministic unlock condition",
        ],
        ["string", "string", "string", "string", "string", "string", "string", "string", "string", "string"],
        [
            ["fisherman", "Fisherman", "building", "fish", "generator", "1", "10", "fish", "1.15", "always"],
            ["boat", "Boat", "building", "fish", "generator", "8", "120", "fish", "1.17", "owned(fisherman) >= 5"],
            ["net", "Net", "building", "fish", "generator", "30", "1800", "fish", "1.2", "owned(boat) >= 3"],
        ],
    )
    write_workbook(
        DATAS / "upgrades.xlsx",
        ["id", "name", "target", "modifier_type", "value", "cost_resource", "base_cost", "unlock_condition"],
        [
            "stable upgrade id",
            "display name",
            "modifier target",
            "modifier type id",
            "modifier value",
            "resource spent",
            "purchase cost",
            "deterministic unlock condition",
        ],
        ["string", "string", "string", "string", "string", "string", "string", "string"],
        [
            ["fisherman_double", "Fisherman Double", "generator:fisherman.output", "multiply", "2", "fish", "500", "owned(fisherman) >= 10"],
            ["boat_flat_bonus", "Boat Flat Bonus", "generator:boat.output", "flat", "5", "fish", "900", "owned(boat) >= 2"],
            ["global_generator_bonus", "Global Generator Bonus", "generator:*.output", "add_pct", "0.5", "fish", "1500", "owned(fisherman) >= 15"],
        ],
    )
    write_workbook(
        DATAS / "constants.xlsx",
        ["id", "value"],
        ["stable constant id", "string-encoded number"],
        ["string", "string"],
        [["starting_fish", "100"], ["starting_prestige_point", "0"]],
    )
    write_workbook(
        DATAS / "milestones.xlsx",
        ["id", "name", "condition", "reward_resource", "reward_amount"],
        ["stable milestone id", "display name", "condition", "resource rewarded", "reward amount"],
        ["string", "string", "string", "string", "string"],
        [
            ["first_boat", "First Boat", "owned(boat) >= 1", "fish", "200"],
            ["small_fleet", "Small Fleet", "owned(boat) >= 3", "fish", "500"],
        ],
    )
    write_workbook(
        DATAS / "prestige_layers.xlsx",
        [
            "id",
            "name",
            "trigger_resource",
            "reward_resource",
            "formula",
            "divisor",
            "exponent",
            "min_gain",
            "reset_resources",
            "unlock_condition",
        ],
        [
            "stable prestige id",
            "display name",
            "resource measured",
            "resource rewarded",
            "YAML formula id",
            "formula divisor",
            "formula exponent",
            "minimum gain",
            "resources reset",
            "condition",
        ],
        ["string", "string", "string", "string", "string", "string", "string", "string", "(list#sep=;),string", "string"],
        [["reef_renown", "Reef Renown", "fish", "prestige_point", "prestige_gain", "600", "0.5", "1", "fish", "owned(boat) >= 5"]],
    )
    export_registered_workbooks(DATAS, EXPORTS)


if __name__ == "__main__":
    main()
